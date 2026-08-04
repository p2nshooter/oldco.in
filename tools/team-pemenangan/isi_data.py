#!/usr/bin/env python3
"""Mengisi sheet DATABASE Excel dari berkas database.json versi web.

Kedua aplikasi memakai satu sumber data yang sama: web/seed/database.json.
Skrip ini menyalinnya ke workbook Excel sehingga isi keduanya identik.

Pemakaian:
    python3 isi_data.py workbook.xlsx web/seed/database.json
"""

from __future__ import annotations

import importlib.util
import json
import pathlib
import sys
import warnings

import openpyxl

warnings.filterwarnings("ignore")

# Rentang baris data pada sheet DATABASE — harus sama dengan FIRST dan LAST di
# build_workbook.py. Dibaca dari sana langsung supaya tidak bisa berbeda.
_spec = importlib.util.spec_from_file_location(
    "build_workbook", pathlib.Path(__file__).with_name("build_workbook.py"))
_bw = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_bw)
FIRST, LAST = _bw.FIRST, _bw.LAST

# kolom Excel -> kunci pada JSON
KOLOM = {
    "B": "nama", "C": "nik", "D": "kk", "E": "jk", "F": "kampung",
    "G": "rt", "H": "rw", "I": "alamat", "J": "namaRt", "K": "namaRw",
    "L": "kadus", "N": "jabatan", "O": "hp", "P": "tps", "Q": "tglLahir",
    "T": "status", "U": "tglGabung", "V": "perekrut", "W": "catatan",
}
# kolom M (KORWIL), R (USIA), S (KELOMPOK USIA) sengaja dilewati — berisi rumus
TANGGAL = {"Q", "U"}


def isi(xlsx: pathlib.Path, data_path: pathlib.Path) -> int:
    data = json.loads(data_path.read_text(encoding="utf-8"))
    anggota = data["members"]
    st = data.get("settings") or {}

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["DATABASE"]

    # Kosongkan dulu seluruh kolom data. Tanpa ini, menjalankan ulang skrip
    # pada berkas yang sudah terisi meninggalkan dua macam sisa: kolom yang
    # sekarang dikosongkan tetap memakai isi lama, dan baris anggota yang sudah
    # dihapus tetap tertinggal di bawah. Kolom rumus (M, R, S, dan kolom bantu)
    # sengaja tidak disentuh.
    for row in range(FIRST, LAST + 1):
        for col in KOLOM:
            ws[f"{col}{row}"].value = None

    for i, m in enumerate(anggota):
        row = FIRST + i
        if row > LAST:
            raise SystemExit(
                f"GAGAL: {len(anggota)} anggota melebihi kapasitas sheet "
                f"({LAST - FIRST + 1} baris). Naikkan ROWS di build_workbook.py.")
        for col, kunci in KOLOM.items():
            nilai = m.get(kunci, "")
            if not nilai:
                continue
            sel = ws[f"{col}{row}"]
            sel.value = nilai                 # tanggal dibiarkan teks apa adanya
            # Nilai yang diawali "=" diperlakukan Excel sebagai rumus, bukan
            # teks. Sebuah alamat yang ditulis "=RUMAH PAK RT" akan berubah
            # menjadi #NAME? dan ikut merusak kolom MASALAH serta seluruh rekap
            # yang membacanya. Jenis selnya dipaksa teks supaya isinya utuh.
            if isinstance(nilai, str) and nilai.startswith("="):
                sel.data_type = "s"

    # daftar pilihan REFERENSI mengikuti berkas yang sama
    ref = wb["REFERENSI"]

    def tulis(col: str, nilai: list, mulai: int = 6):
        for j in range(40):
            sel = ref[f"{col}{mulai + j}"]
            sel.value = nilai[j] if j < len(nilai) else None

    tulis("B", st.get("kampung", []))
    tulis("D", st.get("rt", []))
    tulis("E", st.get("rw", []))
    tulis("F", st.get("jabatan", []))
    tulis("H", st.get("tps", []))
    tulis("O", st.get("namaRt", []))
    tulis("P", st.get("namaRw", []))
    tulis("Q", st.get("kadus", []))
    target = st.get("target", {})
    for j, k in enumerate(st.get("kampung", [])):
        ref[f"C{6 + j}"].value = target.get(k, 0)

    ident = st.get("identitas", {})
    peta_id = [("C49", "calon"), ("C50", "team"), ("C51", "periode"),
               ("C52", "tahun"), ("C53", "desa"), ("C54", "kecamatan"),
               ("C55", "kabupaten"), ("C56", "ketua"), ("C57", "jabatanTtd")]
    for sel, kunci in peta_id:
        nilai = ident.get(kunci, "")
        ref[sel].value = nilai if nilai else "-"

    wb.save(xlsx)
    return len(anggota)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__, file=sys.stderr)
        raise SystemExit(1)
    n = isi(pathlib.Path(sys.argv[1]), pathlib.Path(sys.argv[2]))
    print(f"terisi: {n} anggota ke {sys.argv[1]}")
