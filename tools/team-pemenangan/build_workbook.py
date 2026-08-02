#!/usr/bin/env python3
"""Generator "Aplikasi Team Pemenangan" — versi puncak.

Membangun satu berkas Excel siap pakai untuk pendataan anggota team
pemenangan tingkat desa:

    MENU       dasbor ringkasan + navigasi
    DATABASE   input data anggota (2.000 baris)
    CARI       pencarian cepat + filter berlapis
    REKAP      matriks Kadus x RT, gender, jabatan, status, usia, TPS
    TARGET     target vs realisasi per Kadus
    VALIDASI   daftar baris yang perlu diperbaiki
    CETAK      formulir tanda tangan per korwil
    KARTU      kartu anggota siap potong
    ABSENSI    daftar hadir kegiatan per korwil
    REFERENSI  pengaturan dropdown + identitas + target
    PETUNJUK   manual penggunaan

Pemakaian:
    python3 build_workbook.py [keluaran.xlsx] [--sample]

--sample mengisi 30 baris data contoh; dipakai hanya untuk menguji rumus,
jangan dipakai untuk berkas yang dibagikan.
"""

from __future__ import annotations

import pathlib
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------- konstanta

FIRST = 4                      # baris data pertama pada DATABASE
ROWS = 2000                    # kapasitas baris data
LAST = FIRST + ROWS - 1        # = 2003
DB = "DATABASE"

REF_FIRST, REF_LAST = 6, 45    # rentang daftar pilihan pada REFERENSI
MAX_KADUS = 20                 # baris Kadus yang ditampilkan pada rekap/target
MAX_RT = 20                    # kolom RT pada matriks REKAP
N_CARI = 150                   # baris hasil pada sheet CARI
N_MASALAH = 150                # baris hasil pada sheet VALIDASI
N_CETAK = 25                   # nama per halaman formulir CETAK
N_ABSEN = 25                   # nama per halaman daftar hadir
N_KARTU = 8                    # kartu per halaman (2 kolom x 4 baris)
MENU_KADUS = 12                # baris Kadus pada ringkasan MENU (rekap penuh di REKAP)

FONT = "Arial"

# palet — mengikuti berkas asli (input kuning FFF6D5, aksen merah/navy)
NAVY = "1F3864"
BLUE = "2E75B6"
RED = "C00000"
GREY_BG = "F4F2ED"
GREY_CALC = "F2F2F2"
INPUT = "FFF6D5"
BAND = "EDF3FA"
WHITE = "FFFFFF"
GREEN = "375623"
OK_FILL = "E2EFDA"
WARN_FILL = "FFF2CC"
BAD_FILL = "FFC7CE"

THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="medium", color=NAVY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def f(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill(rgb):
    return PatternFill("solid", fgColor=rgb)


CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

UNLOCKED = Protection(locked=False)


# ------------------------------------------------------------- alat bantu


def put(ws, ref, value, font=None, bg=None, align=None, border=None,
        numfmt=None, unlock=False):
    """Tulis satu sel sekaligus dengan gayanya."""
    c = ws[ref]
    c.value = value
    c.font = font or f()
    if bg:
        c.fill = fill(bg)
    if align:
        c.alignment = align
    if border:
        c.border = border
    if numfmt:
        c.number_format = numfmt
    if unlock:
        c.protection = UNLOCKED
    return c


def title_block(ws, row, judul, sub=None, width="H"):
    """Kop merah + baris navigasi kembali ke MENU."""
    ws.merge_cells(f"A{row}:{width}{row}")
    put(ws, f"A{row}", judul, f(14, True, WHITE), RED, LEFT)
    ws.row_dimensions[row].height = 26
    r = row + 1
    link(ws, f"A{r}", "◀  MENU", "MENU!A1")
    if sub:
        put(ws, f"C{r}", sub, f(9, italic=True, color="595959"), align=LEFT)
    return r


def section(ws, row, text, width="H", color=NAVY):
    ws.merge_cells(f"B{row}:{width}{row}")
    put(ws, f"B{row}", "▌  " + text, f(11, True, WHITE), color, LEFT)
    ws.row_dimensions[row].height = 20


def link(ws, ref, text, location, size=10, bold=True):
    c = put(ws, ref, text, f(size, bold, BLUE), align=LEFT)
    c.hyperlink = Hyperlink(ref=ref, location=location, display=text)
    c.font = Font(name=FONT, size=size, bold=bold, color=BLUE, underline="single")
    return c


def header_row(ws, row, cols, first_col=1, bg=NAVY):
    """cols: list of (judul, lebar_kolom)."""
    for i, (text, width) in enumerate(cols):
        col = get_column_letter(first_col + i)
        put(ws, f"{col}{row}", text, f(10, True, WHITE), bg, CENTER, BOX)
        if width:
            ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 22


def band(ws, row, first_col, last_col, border=BOX):
    """Warna selang-seling + garis untuk satu baris tabel."""
    bg = BAND if row % 2 == 0 else None
    for i in range(first_col, last_col + 1):
        c = ws.cell(row=row, column=i)
        c.border = border
        if bg and not c.fill.fgColor.rgb or bg:
            c.fill = fill(bg) if bg else c.fill


def lock_sheet(ws):
    ws.protection.sheet = True          # tanpa kata sandi: Review > Unprotect Sheet
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


ASSETS = pathlib.Path(__file__).parent / "assets"


def gambar(ws, nama, anchor, lebar, dx=0, dy=0):
    """Sisipkan gambar dari folder assets, diskalakan proporsional ke `lebar` px.

    `dx`/`dy` menggeser gambar dalam piksel dari pojok kiri atas sel jangkar,
    dipakai untuk menengahkan segel pada kartu dan kop surat. Rasio gambar
    selalu dipertahankan agar logo tidak gepeng.

    Dilewati diam-diam bila berkasnya tidak ada, supaya workbook tetap bisa
    dibangun tanpa aset gambar.
    """
    berkas = ASSETS / nama
    if not berkas.exists():
        return None
    img = XLImage(berkas)
    tinggi = int(round(img.height * lebar / img.width))
    img.width, img.height = lebar, tinggi
    if dx or dy:
        col, row = coordinate_to_tuple(anchor)[1] - 1, coordinate_to_tuple(anchor)[0] - 1
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=col, colOff=pixels_to_EMU(dx),
                               row=row, rowOff=pixels_to_EMU(dy)),
            ext=XDRPositiveSize2D(pixels_to_EMU(lebar), pixels_to_EMU(tinggi)))
        ws._images.append(img)
    else:
        ws.add_image(img, anchor)
    return img


def page_print(ws, area=None, landscape=False, fit_width=1, titles=None,
               fit_height=0):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    if area:
        ws.print_area = area
    if titles:
        ws.print_title_rows = titles


# =========================================================== 1. REFERENSI


KAMPUNG_DEFAULT = ["Kp. Gamprit", "Kp. Kuda Kuda", "Kp. Wangkal",
                   "Kp. Tenjolaut", "Kp. Putri Melintang"]
KADUS_DEFAULT = ["Kadus 1", "Kadus 2", "Kadus 3"]
# struktur pengurus desa: (kadus, nama kadus, [(no RK, nama)], [(no RT, nama)])
PENGURUS_DEFAULT = [
    ("Kadus 1", "", [("1", "Maska"), ("2", "Pasni")],
     [("001", "Ikin Lois", "Kp. Tenjolaut"), ("002", "Sarjan", "Kp. Tenjolaut"),
      ("003", "Basir", "Kp. Kuda Kuda"), ("004", "Onin", "Kp. Kuda Kuda"),
      ("005", "Tongkat / Kahidir", "Kp. Kuda Kuda"),
      ("006", "Sa ari", "Kp. Kuda Kuda")]),
    ("Kadus 2", "", [("3", "Sutejo"), ("4", "Ilyas")],
     [("001", "Dedi", "Kp. Gamprit"), ("002", "Rimun", "Kp. Gamprit"),
      ("003", "Adi Ardiansyah", "Kp. Gamprit"), ("004", "Kusnadi", "Kp. Gamprit")]),
    ("Kadus 3", "Markum", [("5", "Uding"), ("6", "Saiman")],
     [("001", "Ropic", "Kp. Gamprit"), ("002", "", "Kp. Gamprit"),
      ("003", "Wandy Suwandi", "Kp. Gamprit"), ("004", "Toyib", "Kp. Wangkal"),
      ("005", "Nemuin", "Kp. Wangkal"), ("006", "Sumintra", "Kp. Gamprit")]),
]
NAMA_RT_DEFAULT = [r[1] for _, _, _, rt in PENGURUS_DEFAULT for r in rt if r[1]]
NAMA_RK_DEFAULT = [n for _, _, rk, _ in PENGURUS_DEFAULT for _, n in rk if n]
RW_DEFAULT = [f"{i:03d}" for i in range(1, 9)]
RT_DEFAULT = [f"{i:03d}" for i in range(1, 16)]
JABATAN_DEFAULT = [
    "Ketua Team", "Wakil Ketua", "Sekretaris", "Bendahara", "Kadus",
    "Ketua RK", "Ketua RT", "Koordinator Wilayah", "Koordinator Lapangan",
    "Anggota", "Relawan",
]
TPS_DEFAULT = [f"{i:02d}" for i in range(1, 21)]
STATUS_DEFAULT = ["Aktif", "Calon", "Nonaktif"]
USIA_DEFAULT = [
    ("Di bawah 17 th", 0),
    ("Pemula 17-22 th", 17),
    ("Muda 23-35 th", 23),
    ("Dewasa 36-50 th", 36),
    ("Matang 51-65 th", 51),
    ("Senior 66 th +", 66),
]
IDENTITAS = [
    ("NAMA CALON", "JONI FAHAMSYAH"),
    ("NAMA TEAM", "TEAM PEMENANGAN"),
    ("PERIODE", "PERIODE 2026 S/D 2034"),
    ("TAHUN", "2026"),
    ("DESA", "SUKAKARYA"),
    ("KECAMATAN", "-"),
    ("KABUPATEN", "-"),
    ("KETUA TEAM", "IDRUS KAMSENO"),
    ("JABATAN PENANDATANGAN", "Ketua Team"),
]


def build_referensi(wb):
    ws = wb.create_sheet("REFERENSI")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLMNOPQRS",
                      [3, 20, 10, 8, 8, 24, 3, 10, 3, 12, 3, 20, 8, 3, 14, 14,
                       14, 3, 46]):
        ws.column_dimensions[col].width = w

    r = title_block(ws, 1, "PENGATURAN  •  DAFTAR PILIHAN, TARGET & IDENTITAS",
                    "Semua isian di sel kuning boleh diubah — dropdown dan kop "
                    "surat ikut berubah otomatis.", width="S")

    hdr = [("B5", "DAFTAR KAMPUNG"), ("C5", "TARGET"), ("D5", "DAFTAR RT"),
           ("E5", "DAFTAR RW"), ("F5", "DAFTAR JABATAN"), ("H5", "DAFTAR TPS"),
           ("J5", "STATUS ANGGOTA"), ("L5", "KELOMPOK USIA"), ("M5", "USIA MIN"),
           ("O5", "NAMA RT"), ("P5", "NAMA RK"), ("Q5", "NAMA KADUS")]
    for ref, text in hdr:
        put(ws, ref, text, f(10, True, WHITE), NAVY, CENTER, BOX)

    def isi(col, values, row0=REF_FIRST, numfmt=None):
        for i in range(REF_LAST - REF_FIRST + 1):
            ref = f"{col}{row0 + i}"
            put(ws, ref, values[i] if i < len(values) else None,
                f(10), INPUT, LEFT, BOX, numfmt, unlock=True)

    isi("B", KAMPUNG_DEFAULT)
    isi("C", [0] * len(KAMPUNG_DEFAULT))
    isi("D", RT_DEFAULT, numfmt="@")
    isi("E", RW_DEFAULT, numfmt="@")
    isi("F", JABATAN_DEFAULT)
    isi("H", TPS_DEFAULT, numfmt="@")
    isi("O", NAMA_RT_DEFAULT)
    isi("P", NAMA_RK_DEFAULT)
    isi("Q", KADUS_DEFAULT)
    for i in range(REF_FIRST, REF_FIRST + 5):
        put(ws, f"J{i}", STATUS_DEFAULT[i - REF_FIRST] if i - REF_FIRST < 3 else None,
            f(10), INPUT, LEFT, BOX, unlock=True)
    for i, (label, umin) in enumerate(USIA_DEFAULT):
        put(ws, f"L{REF_FIRST + i}", label, f(10), INPUT, LEFT, BOX, unlock=True)
        put(ws, f"M{REF_FIRST + i}", umin, f(10), INPUT, CENTER, BOX, unlock=True)

    # --- identitas / kop surat
    put(ws, "B48", "IDENTITAS APLIKASI  (dipakai pada kop MENU, CETAK, KARTU, ABSENSI)",
        f(10, True, WHITE), BLUE, LEFT)
    ws.merge_cells("B48:F48")
    for i, (label, value) in enumerate(IDENTITAS):
        row = 49 + i
        put(ws, f"B{row}", label, f(10, True), GREY_BG, LEFT, BOX)
        ws.merge_cells(f"C{row}:F{row}")
        put(ws, f"C{row}", value, f(10), INPUT, LEFT, BOX, unlock=True)

    # --- struktur pengurus: acuan siapa RT/RK di kadus mana
    put(ws, "B60", "STRUKTUR PENGURUS  •  ACUAN NAMA RT, RK, DAN KADUS",
        f(10, True, WHITE), BLUE, LEFT)
    ws.merge_cells("B60:F60")
    baris = 61
    for kadus, nama_kadus, rk_list, rt_list in PENGURUS_DEFAULT:
        put(ws, f"B{baris}", kadus, f(10, True), GREY_BG, LEFT, BOX)
        put(ws, f"C{baris}", "Kadus", f(9, color="595959"), GREY_BG, LEFT, BOX)
        put(ws, f"D{baris}", nama_kadus or "(belum ada nama)",
            f(10, True) if nama_kadus else f(9, italic=True, color=RED),
            GREY_BG, LEFT, BOX)
        baris += 1
        for no, nm in rk_list:
            put(ws, f"C{baris}", "RK " + no, f(9), None, CENTER, BOX)
            put(ws, f"D{baris}", nm, f(10), None, LEFT, BOX)
            baris += 1
        for no, nm, kp in rt_list:
            put(ws, f"C{baris}", "RT " + no, f(9), None, CENTER, BOX)
            put(ws, f"D{baris}", nm or "(belum ada nama)",
                f(10) if nm else f(9, italic=True, color=RED), None, LEFT, BOX)
            put(ws, f"E{baris}", kp, f(9, color="595959"), None, LEFT, BOX)
            baris += 1
        baris += 1
    put(ws, f"B{baris}", "Nomor RT berulang di tiap kadus dan Kampung Gamprit ada "
        "di Kadus 2 maupun Kadus 3 — nama pengurusnyalah yang membedakan.",
        f(9, italic=True, color="595959"), align=LEFT)

    # --- catatan
    catatan = [
        "CATATAN PENTING",
        "• Sel kuning = boleh diubah. Sel lain berisi rumus.",
        "• TARGET diisi angka jumlah anggota yang ingin dicapai tiap Kadus;",
        "   dipakai pada sheet TARGET untuk menghitung % capaian.",
        "• Menambah pilihan: tulis pada baris kosong tepat di bawah daftar.",
        "  Dropdown bertambah otomatis (kapasitas 40 baris per daftar).",
        "• Menghapus pilihan: kosongkan selnya (Delete), jangan hapus barisnya.",
        "• Jangan menyisipkan baris di tengah daftar — rumus bisa bergeser.",
        "• Bila nama Kadus diganti, data lama pada DATABASE TIDAK ikut berubah.",
        "  Perbaiki lewat filter kolom KADUS di sheet DATABASE.",
        "• USIA MIN harus urut dari kecil ke besar (0, 17, 23, ...).",
        "• KELOMPOK USIA dipakai untuk rekap segmen pemilih.",
    ]
    for i, text in enumerate(catatan):
        ref = f"S{5 + i}"
        put(ws, ref, text, f(10, True) if i == 0 else f(9),
            NAVY if i == 0 else None, LEFT)
        if i == 0:
            ws[ref].font = f(10, True, WHITE)

    ws.freeze_panes = "A6"
    return ws


# ============================================================ 2. DATABASE

DB_COLS = [
    ("NO", 6), ("NAMA LENGKAP", 28), ("NIK KTP (16 digit)", 20),
    ("NO. KK (16 digit)", 20), ("L/P", 6), ("KAMPUNG", 18), ("RT", 7),
    ("RW", 7), ("ALAMAT", 26), ("NAMA RT", 14), ("NAMA RK", 14),
    ("NAMA KADUS", 14), ("KORWIL", 24), ("JABATAN", 20), ("NO. HP", 15),
    ("TPS", 7), ("TGL LAHIR", 13), ("USIA", 7), ("KELOMPOK USIA", 18),
    ("STATUS", 11), ("TGL GABUNG", 13), ("PEREKRUT", 22), ("CATATAN", 30),
]
# kolom bantu (tersembunyi)
HELP_COLS = [("KEY", 24), ("MASALAH", 26), ("IDX_CARI", 10),
             ("IDX_MASALAH", 12), ("KEY_KAMPUNG", 22), ("IDX_CETAK", 10),
             ("IDX_ABSEN", 10), ("IDX_KARTU", 10)]
# A..W data | X..AE bantu
AUTO_COLS = {"A", "M", "R", "S", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"}


def build_database(wb):
    ws = wb.create_sheet(DB)
    ws.sheet_view.showGridLines = False

    r = title_block(ws, 1, "DATABASE ANGGOTA TEAM PEMENANGAN", width="W")
    put(ws, "C2", "Ketik pada baris kosong pertama. Kolom abu-abu (NO, KORWIL, "
        "USIA, KELOMPOK USIA) terisi otomatis — jangan diketik manual.",
        f(9, italic=True, color="595959"), align=LEFT)
    put(ws, "T2", "TOTAL", f(10, True), align=RIGHT)
    put(ws, "U2", f"=COUNTA($B${FIRST}:$B${LAST})", f(11, True, RED), align=LEFT)
    put(ws, "V2", "SISA BARIS", f(10, True), align=RIGHT)
    put(ws, "W2", f"={ROWS}-COUNTA($B${FIRST}:$B${LAST})", f(11, True, GREEN), align=LEFT)

    all_cols = DB_COLS + HELP_COLS
    header_row(ws, 3, all_cols)

    for row in range(FIRST, LAST + 1):
        p = row - 1  # baris sebelumnya
        vals = {
            "A": f'=IF($B{row}="","",COUNTA($B${FIRST}:$B{row}))',
            # korwil: kampung + RT bila nomornya ada
            "M": (f'=IF($F{row}="","",$F{row}&IF($G{row}="",""," - RT "&$G{row}'
                  f'&IF($H{row}="","","/"&$H{row})))'),
            "R": f'=IF(OR($B{row}="",$Q{row}=""),"",DATEDIF($Q{row},TODAY(),"Y"))',
            "S": f'=IF($R{row}="","",INDEX(usia_label,MATCH($R{row},usia_min,1)))',
            "X": f'=IF($M{row}="","",$M{row}&"#"&COUNTIF($M${FIRST}:$M{row},$M{row}))',
            "AB": f'=IF($F{row}="","",$F{row}&"#"&COUNTIF($F${FIRST}:$F{row},$F{row}))',
            "Y": (
                f'=IF($B{row}="","",'
                f'IF($C{row}="","NIK kosong",'
                f'IF(LEN($C{row})<>16,"NIK bukan 16 digit",'
                f'IF(COUNTIF($C${FIRST}:$C${LAST},$C{row}&"*")>1,"NIK ganda",'
                f'IF(AND($D{row}<>"",LEN($D{row})<>16),"No. KK bukan 16 digit",'
                f'IF(AND($F{row}="",$I{row}=""),"Alamat kosong",'
                f'IF($N{row}="","Jabatan kosong",'
                f'IF($O{row}="","No. HP kosong",'
                f'IF(AND($R{row}<>"",$R{row}<17),"Usia di bawah 17","")))))))))'
            ),
            "Z": (
                # kata kunci kosong = cocokkan semua; SEARCH("") menghasilkan
                # #VALUE! di LibreOffice sehingga harus dijaga lebih dulu.
                f'=IF($B{row}="","",IF(AND('
                f'OR(CARI!$C$5="",ISNUMBER(SEARCH(CARI!$C$5,'
                f'$B{row}&" "&$C{row}&" "&$D{row}&" "&$O{row}&" "&$I{row}&" "'
                f'&$F{row}&" "&$J{row}&" "&$K{row}&" "&$L{row}&" "&$V{row}))),'
                f'OR(CARI!$C$6="",$F{row}=CARI!$C$6),'
                f'OR(CARI!$C$7="",$G{row}=CARI!$C$7),'
                f'OR(CARI!$C$8="",$T{row}=CARI!$C$8)),'
                f'MAX($Z$3:Z{p})+1,""))'
            ),
            "AA": f'=IF($Y{row}="","",MAX($AA$3:AA{p})+1)',
            # nomor urut hasil penyaringan untuk tiap sheet cetak; filter yang
            # dikosongkan berarti "semua", sehingga satu rumus melayani cetak
            # semua anggota, per kampung, per RT, per korwil, maupun per TPS
            "AC": (
                f'=IF($B{row}="","",IF(AND('
                f'OR(CETAK!$J$10="",$F{row}=CETAK!$J$10),'
                f'OR(CETAK!$J$11="",$G{row}=CETAK!$J$11),'
                f'OR(CETAK!$J$12="",$P{row}=CETAK!$J$12),'
                f'OR(CETAK!$J$13="",$N{row}=CETAK!$J$13),'
                f'OR(CETAK!$J$14="",$T{row}=CETAK!$J$14)),'
                f'MAX($AC$3:AC{p})+1,""))'
            ),
            "AD": (
                f'=IF($B{row}="","",IF(AND('
                f'OR(ABSENSI!$H$10="",$F{row}=ABSENSI!$H$10),'
                f'OR(ABSENSI!$H$11="",$G{row}=ABSENSI!$H$11),'
                f'OR(ABSENSI!$H$12="",$P{row}=ABSENSI!$H$12)),'
                f'MAX($AD$3:AD{p})+1,""))'
            ),
            "AE": (
                f'=IF($B{row}="","",IF(AND('
                f'OR(KARTU!$K$6="",$F{row}=KARTU!$K$6),'
                f'OR(KARTU!$K$7="",$G{row}=KARTU!$K$7),'
                f'OR(KARTU!$K$8="",$P{row}=KARTU!$K$8)),'
                f'MAX($AE$3:AE{p})+1,""))'
            ),
        }
        for i, (name, _w) in enumerate(all_cols):
            col = get_column_letter(1 + i)
            c = ws.cell(row=row, column=1 + i)
            if col in vals:
                c.value = vals[col]
            c.font = f(10)
            c.border = BOX
            c.alignment = CENTER if col in ("A", "E", "G", "H", "P", "R") else LEFT
            if col in AUTO_COLS:
                c.fill = fill(GREY_CALC)
            if col in ("C", "D", "G", "H", "O", "P"):
                c.number_format = "@"
            if col in ("Q", "U"):
                c.number_format = "dd/mm/yyyy"
                c.alignment = CENTER

    # tabel ber-filter (kolom bantu sengaja di luar tabel)
    tbl = Table(displayName="tblAnggota", ref=f"A3:W{LAST}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

    # sembunyikan kolom bantu
    for i in range(len(DB_COLS), len(all_cols)):
        ws.column_dimensions[get_column_letter(1 + i)].hidden = True

    # ---- validasi data
    def dv(formula, cols, allow_blank=True, kind="list", **kw):
        d = DataValidation(type=kind, formula1=formula, allow_blank=allow_blank, **kw)
        ws.add_data_validation(d)
        for col in cols:
            d.add(f"{col}{FIRST}:{col}{LAST}")
        return d

    dv('"L,P"', ["E"])
    dv("=daftar_kampung", ["F"])
    dv("=daftar_rt", ["G"])
    dv("=daftar_rw", ["H"])
    dv("=daftar_nama_rt", ["J"])
    dv("=daftar_nama_rk", ["K"])
    dv("=daftar_kadus", ["L"])
    dv("=daftar_jabatan", ["N"])
    dv("=daftar_tps", ["P"])
    dv("=daftar_status", ["T"])
    dv("=nama_anggota", ["V"])

    nik = DataValidation(
        type="custom", formula1=f'=OR($C{FIRST}="",AND(LEN($C{FIRST})=16,ISNUMBER($C{FIRST}*1)))',
        allow_blank=True, showErrorMessage=True, errorStyle="warning",
        errorTitle="Periksa NIK",
        error="NIK sebaiknya 16 angka tanpa spasi. Tekan Ya bila memang ingin "
              "menyimpan apa adanya.")
    ws.add_data_validation(nik)
    nik.add(f"C{FIRST}:D{LAST}")

    tgl = DataValidation(
        type="date", operator="between", formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)", allow_blank=True, showErrorMessage=True,
        errorStyle="warning", errorTitle="Format tanggal",
        error="Isi tanggal dengan format hh/bb/tttt, contoh 12/03/1980.")
    ws.add_data_validation(tgl)
    tgl.add(f"Q{FIRST}:Q{LAST}")
    tgl.add(f"U{FIRST}:U{LAST}")

    # ---- pewarnaan otomatis
    # memakai kolom bantu $T (MASALAH) supaya ringan: satu COUNTIF per baris,
    # bukan satu COUNTIF per sel yang diwarnai.
    rng = f"A{FIRST}:W{LAST}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'$Y{FIRST}="NIK ganda"'], fill=fill(BAD_FILL), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'OR($Y{FIRST}="NIK bukan 16 digit",$Y{FIRST}="NIK kosong")'],
        fill=fill(WARN_FILL), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'$T{FIRST}="Nonaktif"'],
        font=Font(name=FONT, size=10, color="808080", italic=True), stopIfTrue=False))

    ws.freeze_panes = "B4"
    ws.sheet_view.zoomScale = 100
    # Siap cetak hasil penyaringan: Excel tidak mencetak baris yang disembunyikan
    # AutoFilter, jadi "cetak semua per TPS / RT / Kadus" cukup dengan menyaring
    # kolomnya lalu Ctrl+P — tanpa baris kosong, berapa pun jumlah datanya.
    page_print(ws, area=f"A3:W{LAST}", landscape=True, titles="3:3")
    ws.oddHeader.center.text = "&\"Arial,Bold\"&12DAFTAR ANGGOTA TEAM PEMENANGAN"
    ws.oddHeader.right.text = "&\"Arial,Italic\"&9&D"
    ws.oddFooter.right.text = "&\"Arial,Regular\"&9Halaman &P dari &N"
    return ws


# ================================================================ 3. MENU


def build_menu(wb):
    ws = wb.create_sheet("MENU", 0)
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [2, 38, 3, 3, 30, 12, 12, 3, 26, 10]):
        ws.column_dimensions[col].width = w

    put(ws, "B2", "=id_team", f(20, True, RED), align=LEFT)
    put(ws, "B3", "=id_calon", f(16, True, NAVY), align=LEFT)
    put(ws, "B4", "=id_periode", f(11, True, "595959"), align=LEFT)
    put(ws, "B6", '="APLIKASI DATABASE ANGGOTA   •   DESA "&id_desa',
        f(11, True, WHITE), NAVY, LEFT)
    ws.merge_cells("B6:G6")
    put(ws, "I2", '="Hari ini: "&TEXT(TODAY(),"dd/mm/yyyy")',
        f(9, italic=True, color="595959"), align=RIGHT)
    # emblem di pojok kanan atas, sejajar blok judul
    gambar(ws, "logo-emblem.png", "I3", 150)

    # ---- panel navigasi
    section(ws, 8, "MENU UTAMA", width="C")
    menu = [
        ("▶   INPUT / DATABASE ANGGOTA", "DATABASE!A1", "Tambah & ubah data anggota"),
        ("▶   CARI DATA ANGGOTA", "CARI!A1", "Cari nama / NIK / no. HP"),
        ("▶   REKAP & ANALISA", "REKAP!A1", "Matriks Kadus x RT, gender, usia"),
        ("▶   TARGET & CAPAIAN", "TARGET!A1", "Target vs realisasi per Kadus"),
        ("▶   VALIDASI DATA", "VALIDASI!A1", "Daftar data yang perlu diperbaiki"),
        ("▶   CETAK DAFTAR / FORMULIR", "CETAK!A1",
         "Semua, per Kadus, RT, korwil, atau TPS"),
        ("▶   CETAK KARTU ANGGOTA", "KARTU!A1",
         "8 per halaman, bisa disaring per wilayah"),
        ("▶   CETAK DAFTAR HADIR", "ABSENSI!A1",
         "Absensi kegiatan, filter sama seperti CETAK"),
        ("▶   PROFIL & MEDIA KAMPANYE", "PROFIL!A1", "Poster dan emblem siap cetak"),
        ("▶   PENGATURAN KADUS / RT / JABATAN", "REFERENSI!A1", "Dropdown, target, identitas"),
        ("▶   PETUNJUK PENGGUNAAN", "PETUNJUK!A1", "Manual + kode makro opsional"),
    ]
    for i, (text, loc, ket) in enumerate(menu):
        row = 9 + i
        link(ws, f"B{row}", text, loc)
        ws[f"B{row}"].fill = fill(BAND if i % 2 == 0 else WHITE)
        ws[f"B{row}"].border = BOX
        put(ws, f"C{row}", "", None, BAND if i % 2 == 0 else WHITE)
        ws.row_dimensions[row].height = 19

    # ---- KPI
    ws.merge_cells("E8:G8")
    put(ws, "E8", "▌  RINGKASAN DATA", f(11, True, WHITE), NAVY, LEFT)

    kpi = [
        ("Total anggota terdata", f'=COUNTA({DB}!$B${FIRST}:$B${LAST})', "#,##0"),
        ("Anggota berstatus Aktif", f'=COUNTIF({DB}!$T${FIRST}:$T${LAST},"Aktif")', "#,##0"),
        ("Laki-laki  /  Perempuan",
         f'=COUNTIF({DB}!$E${FIRST}:$E${LAST},"L")&"  /  "'
         f'&COUNTIF({DB}!$E${FIRST}:$E${LAST},"P")', "@"),
        ("Jumlah kampung aktif",
         f'=SUMPRODUCT((COUNTIF({DB}!$F${FIRST}:$F${LAST},'
         f'REFERENSI!$B${REF_FIRST}:$B${REF_LAST})>0)*'
         f'(REFERENSI!$B${REF_FIRST}:$B${REF_LAST}<>""))', "#,##0"),
        ("Jumlah RT tercakup",
         f'=SUMPRODUCT((COUNTIF({DB}!$G${FIRST}:$G${LAST},'
         f'REFERENSI!$D${REF_FIRST}:$D${REF_LAST})>0)*'
         f'(REFERENSI!$D${REF_FIRST}:$D${REF_LAST}<>""))', "#,##0"),
        ("Rata-rata usia (tahun)",
         f'=IFERROR(ROUND(AVERAGE({DB}!$R${FIRST}:$R${LAST}),0),"-")', "#,##0"),
        ("Total target (semua Kadus)",
         f'=SUM(REFERENSI!$C${REF_FIRST}:$C${REF_LAST})', "#,##0"),
        ("Capaian terhadap target",
         f'=IFERROR(COUNTA({DB}!$B${FIRST}:$B${LAST})/'
         f'SUM(REFERENSI!$C${REF_FIRST}:$C${REF_LAST}),"-")', "0.0%"),
        ("Data perlu diperbaiki", f'=COUNT({DB}!$AA${FIRST}:$AA${LAST})', "#,##0"),
        ("Sisa kapasitas baris", f'={ROWS}-COUNTA({DB}!$B${FIRST}:$B${LAST})', "#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(kpi):
        row = 9 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"E{row}", label, f(10), bg, LEFT, BOX)
        ws.merge_cells(f"F{row}:G{row}")
        put(ws, f"F{row}", formula, f(11, True, NAVY), bg, CENTER, BOX, fmt)

    ws.conditional_formatting.add("F17:G17", CellIsRule(
        operator="greaterThan", formula=["0"], fill=fill(BAD_FILL),
        font=Font(name=FONT, size=11, bold=True, color=RED)))

    # ---- jumlah anggota per Kadus
    section(ws, 21, "JUMLAH ANGGOTA PER KAMPUNG", width="G")
    ws.merge_cells("B21:G21")
    header_row(ws, 22, [("KAMPUNG", None), ("", None), ("", None), ("JUMLAH", None),
                        ("TARGET", None), ("CAPAIAN", None)], first_col=2)
    ws.merge_cells("B22:D22")
    for i in range(MENU_KADUS):
        row = 23 + i
        n = i + 1
        bg = BAND if i % 2 == 0 else WHITE
        ws.merge_cells(f"B{row}:D{row}")
        put(ws, f"B{row}", f'=IFERROR(INDEX(daftar_kampung,{n}),"")', f(10), bg, LEFT, BOX)
        put(ws, f"E{row}", f'=IF($B{row}="","",COUNTIF({DB}!$F${FIRST}:$F${LAST},$B{row}))',
            f(10, True), bg, CENTER, BOX, "#,##0")
        put(ws, f"F{row}", f'=IF($B{row}="","",IFERROR(INDEX(REFERENSI!$C${REF_FIRST}:'
            f'$C${REF_LAST},{n}),0))', f(10), bg, CENTER, BOX, "#,##0")
        put(ws, f"G{row}", f'=IF(OR($B{row}="",$F{row}=0),"",$E{row}/$F{row})',
            f(10, True, GREEN), bg, CENTER, BOX, "0.0%")
    total = 23 + MENU_KADUS
    ws.merge_cells(f"B{total}:D{total}")
    put(ws, f"B{total}", "TOTAL SEMUA KAMPUNG", f(10, True, WHITE), NAVY, LEFT, BOX)
    # TOTAL dihitung dari seluruh data, bukan hanya 12 baris yang tampil di atas,
    # sehingga tetap benar bila jumlah Kadus lebih dari 12.
    put(ws, f"E{total}", f'=COUNTA({DB}!$B${FIRST}:$B${LAST})',
        f(10, True, WHITE), NAVY, CENTER, BOX, "#,##0")
    put(ws, f"F{total}", f"=SUM(REFERENSI!$C${REF_FIRST}:$C${REF_LAST})",
        f(10, True, WHITE), NAVY, CENTER, BOX, "#,##0")
    put(ws, f"G{total}", f'=IF($F{total}=0,"",$E{total}/$F{total})',
        f(10, True, WHITE), NAVY, CENTER, BOX, "0.0%")

    ws.conditional_formatting.add(f"G23:G{total - 1}", CellIsRule(
        operator="greaterThanOrEqual", formula=["1"], fill=fill(OK_FILL)))
    ws.conditional_formatting.add(f"G23:G{total - 1}", CellIsRule(
        operator="lessThan", formula=["0.5"], fill=fill(WARN_FILL)))

    put(ws, f"B{total + 1}",
        f"Menampilkan {MENU_KADUS} kampung pertama — rekap lengkap ada di sheet REKAP "
        "dan TARGET.", f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{total + 1}:G{total + 1}")

    # ---- catatan alur kerja
    foot = total + 3
    put(ws, "B" + str(foot),
        "Alur kerja:  PENGATURAN  →  DATABASE (input)  →  VALIDASI (bersihkan)  →  "
        "REKAP / TARGET (pantau)  →  CETAK / KARTU / ABSENSI (formulir)",
        f(10, True, NAVY), align=LEFT)
    ws.merge_cells(f"B{foot}:I{foot}")
    put(ws, f"B{foot + 1}",
        "Seluruh angka pada halaman ini dihitung otomatis dari sheet DATABASE. "
        "Simpan dengan Ctrl+S setiap selesai menginput.",
        f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{foot + 1}:I{foot + 1}")
    put(ws, f"B{foot + 2}",
        "Sel berlatar kuning = boleh diisi/diubah.  Sel abu-abu = rumus otomatis, "
        "jangan diketik manual.", f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{foot + 2}:I{foot + 2}")

    page_print(ws, area=f"A1:I{foot + 2}", landscape=True)
    return ws


# =============================================================== 4. REKAP


def build_rekap(wb):
    ws = wb.create_sheet("REKAP")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for i in range(MAX_RT + 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 7.5

    last_rt_col = get_column_letter(2 + MAX_RT + 1)   # kolom TOTAL matriks
    title_block(ws, 1, "REKAP  •  ANALISA DATA ANGGOTA", width=last_rt_col)
    put(ws, "C2", "Seluruh angka dihitung otomatis dari sheet DATABASE.",
        f(9, italic=True, color="595959"), align=LEFT)

    row = 4
    # ---------- A. matriks Kadus x RT
    section(ws, row, "A.  MATRIKS JUMLAH ANGGOTA PER KAMPUNG & RT", width=last_rt_col)
    row += 1
    head = row
    put(ws, f"B{head}", "KAMPUNG", f(10, True, WHITE), BLUE, CENTER, BOX)
    for i in range(MAX_RT):
        col = get_column_letter(3 + i)
        put(ws, f"{col}{head}", f"=REFERENSI!$D${REF_FIRST + i}",
            f(9, True, WHITE), BLUE, CENTER, BOX)
    put(ws, f"{last_rt_col}{head}", "TOTAL", f(10, True, WHITE), NAVY, CENTER, BOX)

    first = head + 1
    for k in range(MAX_KADUS):
        r = first + k
        bg = BAND if k % 2 == 0 else WHITE
        put(ws, f"B{r}", f'=IFERROR(INDEX(daftar_kampung,{k + 1}),"")', f(10), bg, LEFT, BOX)
        for i in range(MAX_RT):
            col = get_column_letter(3 + i)
            put(ws, f"{col}{r}",
                f'=IF(OR($B{r}="",{col}${head}=""),"",'
                f'COUNTIFS({DB}!$F${FIRST}:$F${LAST},$B{r},'
                f'{DB}!$G${FIRST}:$G${LAST},{col}${head}))',
                f(9), bg, CENTER, BOX)
        put(ws, f"{last_rt_col}{r}",
            f'=IF($B{r}="","",COUNTIF({DB}!$F${FIRST}:$F${LAST},$B{r}))',
            f(10, True), bg, CENTER, BOX)
    tot = first + MAX_KADUS
    put(ws, f"B{tot}", "TOTAL", f(10, True, WHITE), NAVY, LEFT, BOX)
    for i in range(MAX_RT + 1):
        col = get_column_letter(3 + i)
        put(ws, f"{col}{tot}", f"=SUM({col}{first}:{col}{tot - 1})",
            f(10, True, WHITE), NAVY, CENTER, BOX)
    matrix_first, matrix_tot = first, tot
    row = tot + 2

    # ---------- B. rekap gender + status per Kadus
    section(ws, row, "B.  GENDER & STATUS PER KAMPUNG", width="K")
    row += 1
    hb = row
    for ref, text in [("B", "KAMPUNG"), ("C", "LAKI-LAKI"), ("D", "PEREMPUAN"),
                      ("E", "AKTIF"), ("F", "CALON"), ("G", "NONAKTIF"),
                      ("H", "TOTAL"), ("I", "% DARI TOTAL")]:
        put(ws, f"{ref}{hb}", text, f(9, True, WHITE), BLUE, CENTER, BOX)
    ws.column_dimensions["I"].width = 12
    for k in range(MAX_KADUS):
        r = hb + 1 + k
        bg = BAND if k % 2 == 0 else WHITE
        put(ws, f"B{r}", f'=IFERROR(INDEX(daftar_kampung,{k + 1}),"")', f(10), bg, LEFT, BOX)
        pairs = [("C", "$D$", "L"), ("D", "$D$", "P"), ("E", "$O$", "Aktif"),
                 ("F", "$O$", "Calon"), ("G", "$O$", "Nonaktif")]
        for col, colref, crit in pairs:
            cr = colref.replace("$", "")
            put(ws, f"{col}{r}",
                f'=IF($B{r}="","",COUNTIFS({DB}!$F${FIRST}:$F${LAST},$B{r},'
                f'{DB}!${cr}${FIRST}:${cr}${LAST},"{crit}"))', f(9), bg, CENTER, BOX)
        put(ws, f"H{r}", f'=IF($B{r}="","",COUNTIF({DB}!$F${FIRST}:$F${LAST},$B{r}))',
            f(10, True), bg, CENTER, BOX)
        put(ws, f"I{r}",
            f'=IF(OR($B{r}="",COUNTA({DB}!$B${FIRST}:$B${LAST})=0),"",'
            f'$H{r}/COUNTA({DB}!$B${FIRST}:$B${LAST}))', f(9), bg, CENTER, BOX, "0.0%")
    rb = hb + 1 + MAX_KADUS
    put(ws, f"B{rb}", "TOTAL", f(10, True, WHITE), NAVY, LEFT, BOX)
    for col in "CDEFGH":
        put(ws, f"{col}{rb}", f"=SUM({col}{hb + 1}:{col}{rb - 1})",
            f(10, True, WHITE), NAVY, CENTER, BOX)
    put(ws, f"I{rb}", f'=IF(SUM(H{hb + 1}:H{rb - 1})=0,"",1)',
        f(10, True, WHITE), NAVY, CENTER, BOX, "0.0%")
    gender_head, gender_first = hb, hb + 1
    row = rb + 2

    # ---------- C. rekap per jabatan
    section(ws, row, "C.  REKAP PER JABATAN", width="K")
    row += 1
    hj = row
    for ref, text in [("B", "JABATAN"), ("C", "JUMLAH"), ("D", "AKTIF"), ("E", "% TOTAL")]:
        put(ws, f"{ref}{hj}", text, f(9, True, WHITE), BLUE, CENTER, BOX)
    for i in range(12):
        r = hj + 1 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", f'=IFERROR(INDEX(daftar_jabatan,{i + 1}),"")', f(10), bg, LEFT, BOX)
        put(ws, f"C{r}", f'=IF($B{r}="","",COUNTIF({DB}!$N${FIRST}:$N${LAST},$B{r}))',
            f(10, True), bg, CENTER, BOX)
        put(ws, f"D{r}", f'=IF($B{r}="","",COUNTIFS({DB}!$N${FIRST}:$N${LAST},$B{r},'
            f'{DB}!$T${FIRST}:$T${LAST},"Aktif"))', f(9), bg, CENTER, BOX)
        put(ws, f"E{r}", f'=IF(OR($B{r}="",COUNTA({DB}!$B${FIRST}:$B${LAST})=0),"",'
            f'$C{r}/COUNTA({DB}!$B${FIRST}:$B${LAST}))', f(9), bg, CENTER, BOX, "0.0%")
    row = hj + 14

    # ---------- D. rekap kelompok usia
    section(ws, row, "D.  SEGMEN USIA PEMILIH", width="K")
    row += 1
    hu = row
    for ref, text in [("B", "KELOMPOK USIA"), ("C", "JUMLAH"), ("D", "LAKI-LAKI"),
                      ("E", "PEREMPUAN"), ("F", "% TOTAL")]:
        put(ws, f"{ref}{hu}", text, f(9, True, WHITE), BLUE, CENTER, BOX)
    for i in range(len(USIA_DEFAULT)):
        r = hu + 1 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", f"=REFERENSI!$L${REF_FIRST + i}", f(10), bg, LEFT, BOX)
        put(ws, f"C{r}", f'=IF($B{r}="","",COUNTIF({DB}!$S${FIRST}:$S${LAST},$B{r}))',
            f(10, True), bg, CENTER, BOX)
        put(ws, f"D{r}", f'=IF($B{r}="","",COUNTIFS({DB}!$S${FIRST}:$S${LAST},$B{r},'
            f'{DB}!$E${FIRST}:$E${LAST},"L"))', f(9), bg, CENTER, BOX)
        put(ws, f"E{r}", f'=IF($B{r}="","",COUNTIFS({DB}!$S${FIRST}:$S${LAST},$B{r},'
            f'{DB}!$E${FIRST}:$E${LAST},"P"))', f(9), bg, CENTER, BOX)
        put(ws, f"F{r}", f'=IF(OR($B{r}="",COUNT({DB}!$R${FIRST}:$R${LAST})=0),"",'
            f'$C{r}/COUNT({DB}!$R${FIRST}:$R${LAST}))', f(9), bg, CENTER, BOX, "0.0%")
    put(ws, f"B{hu + 1 + len(USIA_DEFAULT)}", "Tanpa tanggal lahir", f(10, italic=True),
        WARN_FILL, LEFT, BOX)
    put(ws, f"C{hu + 1 + len(USIA_DEFAULT)}",
        f'=COUNTA({DB}!$B${FIRST}:$B${LAST})-COUNT({DB}!$R${FIRST}:$R${LAST})',
        f(10, True), WARN_FILL, CENTER, BOX)
    row = hu + len(USIA_DEFAULT) + 3

    # ---------- E. rekap per TPS
    section(ws, row, "E.  SEBARAN PER TPS", width="K")
    row += 1
    ht = row
    put(ws, f"B{ht}", "TPS", f(9, True, WHITE), BLUE, CENTER, BOX)
    for i in range(10):
        col = get_column_letter(3 + i)
        put(ws, f"{col}{ht}", f"=REFERENSI!$H${REF_FIRST + i}",
            f(9, True, WHITE), BLUE, CENTER, BOX)
    put(ws, f"B{ht + 1}", "JUMLAH", f(10, True), GREY_BG, CENTER, BOX)
    for i in range(10):
        col = get_column_letter(3 + i)
        put(ws, f"{col}{ht + 1}",
            f'=IF({col}${ht}="","",COUNTIF({DB}!$P${FIRST}:$P${LAST},{col}${ht}))',
            f(10, True), None, CENTER, BOX)
    row = ht + 3

    # ---------- F. daftar anggota per korwil
    section(ws, row, "F.  DAFTAR ANGGOTA PER KORWIL", width="K")
    row += 1
    sel = row
    put(ws, f"B{sel}", "PILIH KAMPUNG", f(10, True), align=LEFT)
    put(ws, f"D{sel}", None, f(10, True, NAVY), INPUT, CENTER, BOX, unlock=True)
    put(ws, f"E{sel}", "RT", f(10, True), align=RIGHT)
    put(ws, f"F{sel}", None, f(10, True, NAVY), INPUT, CENTER, BOX, "@", unlock=True)
    put(ws, f"G{sel}", "RW", f(10, True), align=RIGHT)
    put(ws, f"H{sel}", None, f(10, True, NAVY), INPUT, CENTER, BOX, "@", unlock=True)
    put(ws, f"I{sel}", "JUMLAH", f(10, True), align=RIGHT)
    put(ws, f"J{sel}", f'=COUNTIF({DB}!$M${FIRST}:$M${LAST},$AA$1)',
        f(11, True, RED), None, CENTER, BOX, "#,##0")
    put(ws, f"K{sel}", "◀  isi sel kuning", f(9, italic=True, color="595959"),
        align=LEFT)
    # Kunci korwil disimpan di sel bantu AA1. Bentuknya harus persis sama
    # dengan kolom KORWIL pada DATABASE, termasuk bagian RW-nya.
    put(ws, "AA1",
        f'=IF(OR($D${sel}="",$F${sel}=""),"~",$D${sel}&" - RT "&$F${sel}'
        f'&IF($H${sel}="","","/"&$H${sel}))', f(9))
    ws.column_dimensions["AA"].hidden = True

    dvk = DataValidation(type="list", formula1="=daftar_kampung", allow_blank=True)
    dvr = DataValidation(type="list", formula1="=daftar_rt", allow_blank=True)
    dvw = DataValidation(type="list", formula1="=daftar_rw", allow_blank=True)
    for d, ref in ((dvk, f"D{sel}"), (dvr, f"F{sel}"), (dvw, f"H{sel}")):
        ws.add_data_validation(d)
        d.add(ref)

    hd = sel + 1
    cols = [("NO", 6), ("NAMA LENGKAP", 28), ("NIK", 20), ("L/P", 6), ("JABATAN", 20),
            ("NO. HP", 15), ("STATUS", 10), ("ALAMAT", 26)]
    for i, (text, w) in enumerate(cols):
        col = get_column_letter(2 + i)
        put(ws, f"{col}{hd}", text, f(9, True, WHITE), NAVY, CENTER, BOX)
    src = {"C": "B", "D": "C", "E": "E", "F": "N", "G": "O", "H": "T", "I": "I"}
    for i in range(40):
        r = hd + 1 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", i + 1, f(9), bg, CENTER, BOX)
        for col, dbcol in src.items():
            put(ws, f"{col}{r}",
                f'=IFERROR(INDEX({DB}!${dbcol}${FIRST}:${dbcol}${LAST},'
                f'MATCH($AA$1&"#"&$B{r},{DB}!$X${FIRST}:$X${LAST},0))&"","")',
                f(9), bg, LEFT if col in ("C", "I", "F") else CENTER, BOX,
                "@" if col in ("D", "G") else None)
    last_row = hd + 41

    put(ws, f"B{last_row + 1}",
        "Butuh pencarian bebas (nama / NIK / no. HP)? Gunakan sheet CARI.",
        f(9, italic=True, color="595959"), align=LEFT)

    # ---------- grafik
    chart = BarChart()
    chart.type = "col"
    chart.title = "Jumlah Anggota per Kadus"
    chart.height, chart.width = 7.5, 16
    data = Reference(ws, min_col=3 + MAX_RT, min_row=matrix_first,
                     max_row=matrix_tot - 1)
    cats = Reference(ws, min_col=2, min_row=matrix_first, max_row=matrix_tot - 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"B{last_row + 3}")

    pie = PieChart()
    pie.title = "Komposisi Laki-laki / Perempuan"
    pie.height, pie.width = 7.5, 10
    pdata = Reference(ws, min_col=3, max_col=4, min_row=rb)
    plabel = Reference(ws, min_col=3, max_col=4, min_row=gender_head)
    pie.add_data(pdata, from_rows=True, titles_from_data=False)
    pie.set_categories(plabel)
    ws.add_chart(pie, f"L{last_row + 3}")

    ws.freeze_panes = "B5"
    lock_sheet(ws)
    page_print(ws, area=f"A1:{last_rt_col}{last_row + 1}", landscape=True)
    return ws


# ============================================================== 5. TARGET


def build_target(wb):
    ws = wb.create_sheet("TARGET")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [2, 24, 12, 12, 12, 12, 12, 30, 3]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "TARGET & CAPAIAN PER KAMPUNG")
    put(ws, "C2", "Ubah angka target pada sheet PENGATURAN kolom TARGET.",
        f(9, italic=True, color="595959"), align=LEFT)

    section(ws, 4, "PERBANDINGAN TARGET DENGAN REALISASI")
    hd = 5
    for ref, text in [("B", "KAMPUNG"), ("C", "TARGET"), ("D", "REALISASI"),
                      ("E", "AKTIF"), ("F", "KURANG"), ("G", "% CAPAIAN"),
                      ("H", "GRAFIK CAPAIAN")]:
        put(ws, f"{ref}{hd}", text, f(10, True, WHITE), NAVY, CENTER, BOX)

    for k in range(MAX_KADUS):
        r = hd + 1 + k
        n = k + 1
        bg = BAND if k % 2 == 0 else WHITE
        put(ws, f"B{r}", f'=IFERROR(INDEX(daftar_kampung,{n}),"")', f(10), bg, LEFT, BOX)
        put(ws, f"C{r}", f'=IF($B{r}="","",IFERROR(INDEX(REFERENSI!$C${REF_FIRST}:'
            f'$C${REF_LAST},{n}),0))', f(10), bg, CENTER, BOX, "#,##0")
        put(ws, f"D{r}", f'=IF($B{r}="","",COUNTIF({DB}!$F${FIRST}:$F${LAST},$B{r}))',
            f(10, True), bg, CENTER, BOX, "#,##0")
        put(ws, f"E{r}", f'=IF($B{r}="","",COUNTIFS({DB}!$F${FIRST}:$F${LAST},$B{r},'
            f'{DB}!$T${FIRST}:$T${LAST},"Aktif"))', f(10), bg, CENTER, BOX, "#,##0")
        put(ws, f"F{r}", f'=IF($B{r}="","",MAX(0,$C{r}-$D{r}))', f(10), bg, CENTER,
            BOX, "#,##0")
        put(ws, f"G{r}", f'=IF(OR($B{r}="",$C{r}=0),"",$D{r}/$C{r})',
            f(10, True), bg, CENTER, BOX, "0.0%")
        put(ws, f"H{r}", f'=IF($G{r}="","",REPT("█",ROUND(MIN($G{r},1)*20,0)))',
            f(10, color=GREEN), bg, LEFT, BOX)

    tot = hd + 1 + MAX_KADUS
    put(ws, f"B{tot}", "TOTAL", f(10, True, WHITE), NAVY, LEFT, BOX)
    for col in "CDEF":
        put(ws, f"{col}{tot}", f"=SUM({col}{hd + 1}:{col}{tot - 1})",
            f(10, True, WHITE), NAVY, CENTER, BOX, "#,##0")
    put(ws, f"G{tot}", f'=IF($C{tot}=0,"",$D{tot}/$C{tot})',
        f(10, True, WHITE), NAVY, CENTER, BOX, "0.0%")
    put(ws, f"H{tot}", f'=IF($G{tot}="","",REPT("█",ROUND(MIN($G{tot},1)*20,0)))',
        f(10, True, WHITE), NAVY, LEFT, BOX)

    rng = f"G{hd + 1}:G{tot - 1}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThanOrEqual", formula=["1"], fill=fill(OK_FILL),
        font=Font(name=FONT, size=10, bold=True, color=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["0.5", "0.999999"], fill=fill(WARN_FILL)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0.5"], fill=fill(BAD_FILL),
        font=Font(name=FONT, size=10, bold=True, color=RED)))

    note = tot + 2
    put(ws, f"B{note}", "Keterangan warna:", f(10, True), align=LEFT)
    for i, (text, bg) in enumerate([
            ("Hijau  : capaian sudah 100% atau lebih", OK_FILL),
            ("Kuning : capaian 50% – 99%", WARN_FILL),
            ("Merah  : capaian di bawah 50%", BAD_FILL)]):
        put(ws, f"B{note + 1 + i}", text, f(9), bg, LEFT, BOX)
        ws.merge_cells(f"B{note + 1 + i}:D{note + 1 + i}")
    put(ws, f"B{note + 5}",
        "Angka TARGET adalah asumsi yang diisi pengguna pada sheet PENGATURAN "
        "(kolom TARGET, sebaris dengan nama Kadus) — bukan hasil perhitungan.",
        f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{note + 5}:H{note + 5}")

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Target vs Realisasi per Kadus"
    chart.height, chart.width = 10, 18
    data = Reference(ws, min_col=3, max_col=4, min_row=hd, max_row=tot - 1)
    cats = Reference(ws, min_col=2, min_row=hd + 1, max_row=tot - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{note + 7}")

    ws.freeze_panes = "B6"
    lock_sheet(ws)
    page_print(ws, area=f"A1:H{note + 5}", landscape=True)
    return ws


# ================================================================ 6. CARI


def build_cari(wb):
    ws = wb.create_sheet("CARI")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJK", [2, 18, 26, 20, 8, 16, 8, 20, 15, 11, 28]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "CARI DATA ANGGOTA", width="K")
    put(ws, "C2", "Ketik sebagian nama / NIK / no. HP / alamat pada sel kuning. "
        "Hasil muncul seketika.", f(9, italic=True, color="595959"), align=LEFT)

    section(ws, 4, "KRITERIA PENCARIAN", width="K")
    kriteria = [
        (5, "KATA KUNCI", "ketik sebagian nama, NIK, no. HP, alamat, atau perekrut"),
        (6, "FILTER KADUS", "kosongkan = semua Kadus"),
        (7, "FILTER RT", "kosongkan = semua RT"),
        (8, "FILTER STATUS", "kosongkan = semua status"),
    ]
    for row, label, ket in kriteria:
        put(ws, f"B{row}", label, f(10, True), GREY_BG, LEFT, BOX)
        put(ws, f"C{row}", None, f(11, True, NAVY), INPUT, LEFT, BOX, "@", unlock=True)
        put(ws, f"D{row}", "◀  " + ket, f(9, italic=True, color="595959"), align=LEFT)

    for row, formula in [(6, "=daftar_kampung"), (7, "=daftar_rt"), (8, "=daftar_status")]:
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"C{row}")

    put(ws, "B10", "DITEMUKAN", f(11, True, WHITE), NAVY, CENTER, BOX)
    put(ws, "C10", f'=COUNT({DB}!$Z${FIRST}:$Z${LAST})&" data"',
        f(12, True, RED), None, LEFT, BOX)
    put(ws, "D10", f"Maksimal {N_CARI} baris pertama yang ditampilkan di bawah. "
        "Persempit kata kunci bila hasil terpotong.",
        f(9, italic=True, color="595959"), align=LEFT)

    hd = 12
    cols = [("NO", None), ("NAMA LENGKAP", None), ("NIK", None), ("L/P", None),
            ("KAMPUNG", None), ("RT", None), ("JABATAN", None), ("NO. HP", None),
            ("STATUS", None), ("ALAMAT", None)]
    header_row(ws, hd, cols, first_col=2)
    src = {"C": "B", "D": "C", "E": "E", "F": "F", "G": "G", "H": "N",
           "I": "O", "J": "T", "K": "I"}
    for i in range(N_CARI):
        r = hd + 1 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", i + 1, f(9), bg, CENTER, BOX)
        for col, dbcol in src.items():
            put(ws, f"{col}{r}",
                f'=IFERROR(INDEX({DB}!${dbcol}${FIRST}:${dbcol}${LAST},'
                f'MATCH($B{r},{DB}!$Z${FIRST}:$Z${LAST},0))&"","")',
                f(9), bg, LEFT if col in ("C", "H", "K") else CENTER, BOX,
                "@" if col in ("D", "G", "I") else None)

    ws.freeze_panes = "B13"
    lock_sheet(ws)
    page_print(ws, area=f"A1:K{hd + N_CARI}", landscape=True, titles=f"{hd}:{hd}")
    return ws


# ============================================================ 7. VALIDASI


def build_validasi(wb):
    ws = wb.create_sheet("VALIDASI")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [2, 8, 30, 20, 18, 8, 26, 18]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "VALIDASI  •  DATA YANG PERLU DIPERBAIKI")
    put(ws, "C2", "Perbaiki langsung pada sheet DATABASE; daftar di bawah "
        "menyusut otomatis.", f(9, italic=True, color="595959"), align=LEFT)

    section(ws, 4, "RINGKASAN MASALAH")
    ringkas = [
        ("Total data terdata", f'=COUNTA({DB}!$B${FIRST}:$B${LAST})', False),
        ("Total baris bermasalah", f'=COUNT({DB}!$AA${FIRST}:$AA${LAST})', True),
        ("NIK kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"NIK kosong")', True),
        ("NIK bukan 16 digit", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"NIK bukan 16 digit")', True),
        ("NIK ganda (dobel)", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"NIK ganda")', True),
        ("Kadus kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"Kadus kosong")', True),
        ("RT kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"RT kosong")', True),
        ("Jabatan kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"Jabatan kosong")', True),
        ("No. HP kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"No. HP kosong")', True),
        ("Alamat kosong", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"Alamat kosong")', True),
        ("Usia di bawah 17 tahun", f'=COUNTIF({DB}!$Y${FIRST}:$Y${LAST},"Usia di bawah 17")', True),
        ("Tanpa tanggal lahir",
         f'=COUNTA({DB}!$B${FIRST}:$B${LAST})-COUNT({DB}!$R${FIRST}:$R${LAST})', False),
        ("Tanpa status keanggotaan",
         f'=COUNTIFS({DB}!$B${FIRST}:$B${LAST},"<>",{DB}!$T${FIRST}:$T${LAST},"")', False),
    ]
    for i, (label, formula, warn) in enumerate(ringkas):
        r = 5 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", label, f(10), bg, LEFT, BOX)
        ws.merge_cells(f"B{r}:D{r}")
        put(ws, f"E{r}", formula, f(11, True, NAVY), bg, CENTER, BOX, "#,##0")
        if warn:
            ws.conditional_formatting.add(f"E{r}", CellIsRule(
                operator="greaterThan", formula=["0"], fill=fill(BAD_FILL),
                font=Font(name=FONT, size=11, bold=True, color=RED)))
    after = 5 + len(ringkas) + 1

    put(ws, f"B{after}", "Satu baris hanya menampilkan satu masalah utama "
        "(urutan prioritas: NIK → Kadus → RT → Jabatan → HP → Alamat → usia). "
        "Setelah diperbaiki, masalah berikutnya pada baris yang sama akan muncul.",
        f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{after}:H{after}")

    hd = after + 2
    section(ws, hd - 1, "DAFTAR BARIS YANG PERLU DIPERBAIKI")
    header_row(ws, hd, [("NO", None), ("NAMA LENGKAP", None), ("NIK", None),
                        ("KAMPUNG", None), ("RT", None), ("MASALAH", None),
                        ("NO. HP", None)], first_col=2)
    src = {"C": "B", "D": "C", "E": "F", "F": "G", "G": "Y", "H": "O"}
    for i in range(N_MASALAH):
        r = hd + 1 + i
        bg = BAND if i % 2 == 0 else WHITE
        put(ws, f"B{r}", i + 1, f(9), bg, CENTER, BOX)
        for col, dbcol in src.items():
            put(ws, f"{col}{r}",
                f'=IFERROR(INDEX({DB}!${dbcol}${FIRST}:${dbcol}${LAST},'
                f'MATCH($B{r},{DB}!$AA${FIRST}:$AA${LAST},0))&"","")',
                f(9, bold=(col == "G"), color=RED if col == "G" else "000000"),
                bg, LEFT if col in ("C", "G") else CENTER, BOX,
                "@" if col in ("D", "F") else None)

    ws.freeze_panes = "B" + str(hd + 1)
    lock_sheet(ws)
    page_print(ws, area=f"A1:H{hd + N_MASALAH}", landscape=True, titles=f"{hd}:{hd}")
    return ws


# =============================================================== 8. CETAK


def build_cetak(wb):
    """Formulir cetak serbaguna.

    Lima filter opsional (Kadus, RT, TPS, jabatan, status) menyaring daftar.
    Filter yang dikosongkan berarti "semua", jadi satu sheet ini melayani cetak
    seluruh anggota, per Kadus, per RT, per korwil (Kadus+RT), maupun per TPS.
    Judul dan jumlah pada kop menyesuaikan sendiri.
    """
    ws = wb.create_sheet("CETAK")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLM",
                      [4, 24, 18, 18, 24, 15, 13, 14, 3, 19, 17, 3, 9]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["M"].hidden = True     # penunjuk baris DATABASE

    # emblem sebagai kop surat di kiri; teks judul tetap rata tengah A:G
    gambar(ws, "logo-emblem.png", "A2", 88)
    put(ws, "A2", "=id_team", f(16, True), align=CENTER)
    put(ws, "A3", "=id_calon", f(14, True), align=CENTER)
    put(ws, "A4", "=id_periode", f(11), align=CENTER)
    put(ws, "A5", '="DESA "&id_desa&IF(id_kecamatan="-",""," • KECAMATAN "&id_kecamatan)'
        '&IF(id_kabupaten="-",""," • KABUPATEN "&id_kabupaten)', f(10), align=CENTER)
    for r in range(2, 6):
        ws.merge_cells(f"A{r}:H{r}")
    for col in "ABCDEFGH":
        ws[f"{col}6"].border = Border(bottom=MED)

    put(ws, "A8", "DAFTAR ANGGOTA TEAM PEMENANGAN", f(12, True), align=CENTER)
    ws.merge_cells("A8:H8")

    # ---- panel pemilihan (di luar area cetak)
    put(ws, "I9", "PANEL — TIDAK IKUT DICETAK", f(9, True, WHITE), BLUE, CENTER)
    ws.merge_cells("I9:J9")
    panel = [("Pilih Kadus", "=daftar_kampung"), ("Pilih RT", "=daftar_rt"),
             ("Pilih TPS", "=daftar_tps"), ("Pilih Jabatan", "=daftar_jabatan"),
             ("Pilih Status", "=daftar_status")]
    for i, (label, formula) in enumerate(panel):
        r = 10 + i
        put(ws, f"I{r}", label, f(10, True), GREY_BG, LEFT, BOX)
        put(ws, f"J{r}", None, f(10, True, NAVY), INPUT, CENTER, BOX, "@", unlock=True)
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"J{r}")

    put(ws, "I15", "Halaman ke-", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "J15", 1, f(11, True, NAVY), INPUT, CENTER, BOX, unlock=True)
    dvh = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1",
                         showErrorMessage=True, errorStyle="warning",
                         errorTitle="Halaman", error="Isi angka 1 atau lebih.")
    ws.add_data_validation(dvh)
    dvh.add("J15")
    put(ws, "I16", "Jumlah data", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "J16", f'=COUNT({DB}!$AC${FIRST}:$AC${LAST})', f(11, True, RED),
        None, CENTER, BOX, "#,##0")
    put(ws, "I17", "Total halaman", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "J17", f'=MAX(1,ROUNDUP($J$16/{N_CETAK},0))', f(11, True, NAVY),
        None, CENTER, BOX, "#,##0")
    link(ws, "I19", "◀  MENU", "MENU!A1")
    catatan = [
        (20, "Kosongkan semua filter = cetak SELURUH anggota."),
        (21, "Isi satu filter saja untuk cetak per Kadus, per RT, atau per TPS."),
        (22, "Isi Kadus + RT sekaligus untuk cetak per korwil."),
        (23, f"Satu halaman memuat {N_CETAK} nama. Bila Total halaman lebih dari 1,"),
        (24, "ganti angka Halaman ke- lalu Ctrl+P lagi."),
        (25, "Daftar sangat panjang? Saring kolom di sheet DATABASE lalu Ctrl+P —"),
        (26, "baris yang tersembunyi filter tidak ikut tercetak."),
    ]
    for r, teks in catatan:
        put(ws, f"I{r}", teks, f(9, italic=True, color="595959"), align=LEFT)

    # ---- keterangan kop yang menyesuaikan filter
    judul = ('=IF(COUNTA($J$10:$J$14)=0,"SELURUH ANGGOTA",'
             'MID(IF($J$10="",""," • "&$J$10)'
             '&IF($J$11="",""," • RT "&$J$11)'
             '&IF($J$12="",""," • TPS "&$J$12)'
             '&IF($J$13="",""," • "&$J$13)'
             '&IF($J$14="",""," • "&$J$14),4,200))')
    put(ws, "A10", "KELOMPOK", f(11, True), align=LEFT)
    put(ws, "C10", ":", f(11, True), align=CENTER)
    put(ws, "D10", judul, f(11, True, NAVY), align=LEFT)
    ws.merge_cells("D10:H10")
    put(ws, "A11", "JUMLAH", f(11, True), align=LEFT)
    put(ws, "C11", ":", f(11, True), align=CENTER)
    put(ws, "D11", '=$J$16&" orang"&IF($J$17>1,"     (halaman "&$J$15&" dari "'
        '&$J$17&")","")', f(11, True), align=LEFT)
    ws.merge_cells("D11:H11")

    hd = 13
    header_row(ws, hd, [("NO", None), ("NAMA LENGKAP", None), ("NIK KTP", None),
                        ("NO. KK", None), ("ALAMAT", None), ("JABATAN", None),
                        ("NO. HP", None), ("TANDA TANGAN", None)])
    # satu MATCH per baris, sisanya INDEX yang murah
    src = {"B": "B", "C": "C", "D": "D", "E": "M", "F": "N", "G": "O"}
    for i in range(N_CETAK):
        r = hd + 1 + i
        urut = f'(($J$15-1)*{N_CETAK}+{i + 1})'
        put(ws, f"M{r}", f'=IFERROR(MATCH({urut},{DB}!$AC${FIRST}:$AC${LAST},0),"")',
            f(9))
        put(ws, f"A{r}", f'=IF($M{r}="","",{urut})', f(10), None, CENTER, BOX)
        for col, dbcol in src.items():
            put(ws, f"{col}{r}",
                f'=IF($M{r}="","",INDEX({DB}!${dbcol}${FIRST}:${dbcol}${LAST},$M{r})&"")',
                f(10), None, LEFT if col in ("B", "E", "F") else CENTER, BOX,
                "@" if col in ("C", "D", "G") else None)
        put(ws, f"H{r}", f'=IF($M{r}="","",$A{r}&".")', f(9, color="808080"),
            None, LEFT, BOX)
        ws.row_dimensions[r].height = 22

    ttd = hd + N_CETAK + 2
    put(ws, f"F{ttd}", f'=id_desa&", ......................... "&id_tahun',
        f(11), align=CENTER)
    put(ws, f"F{ttd + 1}", "=id_jabatan_ttd", f(11), align=CENTER)
    put(ws, f"F{ttd + 5}", "=id_ketua", f(11, True, color="000000"), align=CENTER)
    for r in (ttd, ttd + 1, ttd + 5):
        ws.merge_cells(f"F{r}:H{r}")
    for col in "FGH":
        ws[f"{col}{ttd + 5}"].border = Border(top=Side(style="thin", color="000000"))

    lock_sheet(ws)
    page_print(ws, area=f"A1:H{ttd + 6}", landscape=True, fit_height=1)
    return ws


# =============================================================== 9. KARTU


def build_kartu(wb):
    ws = wb.create_sheet("KARTU")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJK", [2, 10, 12, 22, 3, 10, 12, 22, 3, 20, 20]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "KARTU ANGGOTA  •  SIAP POTONG", width="H")
    put(ws, "C2", "Isi nomor awal, lalu Ctrl+P. Satu halaman berisi 8 kartu.",
        f(9, italic=True, color="595959"), align=LEFT)

    put(ws, "J4", "PANEL — TIDAK DICETAK", f(9, True, WHITE), BLUE, CENTER)
    ws.merge_cells("J4:K4")
    put(ws, "J5", "Nomor awal", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "K5", 1, f(11, True, NAVY), INPUT, CENTER, BOX, unlock=True)
    for i, (label, formula) in enumerate([("Pilih Kampung", "=daftar_kampung"),
                                          ("Pilih RT", "=daftar_rt"),
                                          ("Pilih TPS", "=daftar_tps")]):
        r = 6 + i
        put(ws, f"J{r}", label, f(10, True), GREY_BG, LEFT, BOX)
        put(ws, f"K{r}", None, f(10, True, NAVY), INPUT, CENTER, BOX, "@", unlock=True)
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        d.add(f"K{r}")
    put(ws, "J9", "Jumlah tersaring", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "K9", f'=COUNT({DB}!$AE${FIRST}:$AE${LAST})', f(11, True, RED),
        None, CENTER, BOX, "#,##0")
    put(ws, "J10", "Halaman berikutnya", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "K10", f'=IF($K$5+{N_KARTU}>$K$9,"— sudah halaman terakhir —",'
        f'"isi "&($K$5+{N_KARTU})&" untuk {N_KARTU} kartu berikutnya")',
        f(9, italic=True), None, LEFT, BOX)
    dvn = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1",
                         showErrorMessage=True, errorStyle="warning",
                         errorTitle="Nomor awal", error="Isi angka 1 atau lebih.")
    ws.add_data_validation(dvn)
    dvn.add("K5")
    link(ws, "J12", "◀  MENU", "MENU!A1")
    put(ws, "J13", "Kosongkan filter = kartu untuk seluruh anggota.",
        f(9, italic=True, color="595959"), align=LEFT)

    # Tiap kartu: bilah judul, lalu badan berisi segel di kiri dan data di
    # kanan — segel dibuat cukup besar (62 px) agar tulisan emblem tetap
    # terbaca; ukuran mini di bilah judul terlihat sebagai noda gelap.
    start_row = 4
    tinggi_baris = 17          # tinggi baris data kartu, dalam poin
    px_per_baris = tinggi_baris * 4 / 3
    segel = 62
    for idx in range(N_KARTU):
        rblock, cblock = idx // 2, idx % 2
        top = start_row + rblock * 8
        cs, cl, cv = ("B", "C", "D") if cblock == 0 else ("F", "G", "H")
        no = f"($K$5+{idx})"
        # nomor mengikuti hasil penyaringan (kolom Z), bukan nomor urut global
        lookup = (f'IFERROR(INDEX({DB}!$%s${FIRST}:$%s${LAST},'
                  f'MATCH({no},{DB}!$AE${FIRST}:$AE${LAST},0))&"","")')

        ws.merge_cells(f"{cs}{top}:{cv}{top}")
        put(ws, f"{cs}{top}", '=id_team&" — "&id_calon', f(9, True, WHITE), NAVY,
            CENTER, BOX)
        ws.row_dimensions[top].height = 20

        rows = [
            ("NAMA", lookup % ("B", "B"), f(10, True)),
            ("NIK", lookup % ("C", "C"), f(9)),
            ("ALAMAT", lookup % ("M", "M"), f(9)),
            ("JABATAN", lookup % ("N", "N"), f(9)),
            ("NO. HP", lookup % ("O", "O"), f(9)),
        ]
        # segel menempati satu blok tergabung di sisi kiri badan kartu
        ws.merge_cells(f"{cs}{top + 1}:{cs}{top + len(rows)}")
        put(ws, f"{cs}{top + 1}", None, f(9), None, CENTER, BOX)
        tinggi_badan = len(rows) * px_per_baris
        gambar(ws, "logo-emblem.png", f"{cs}{top + 1}", segel,
               dx=6, dy=int(round((tinggi_badan - segel) / 2)))

        for j, (label, formula, font) in enumerate(rows):
            r = top + 1 + j
            put(ws, f"{cl}{r}", label, f(8, color="595959"), GREY_BG, LEFT, BOX)
            put(ws, f"{cv}{r}", "=" + formula, font, None, LEFT, BOX,
                "@" if label in ("NIK", "NO. HP") else None)
            ws.row_dimensions[r].height = tinggi_baris
        foot = top + len(rows) + 1
        put(ws, f"{cs}{foot}", f'="No. "&{no}', f(8, color="808080"), None, LEFT)
        ws.merge_cells(f"{cs}{foot}:{cl}{foot}")
        put(ws, f"{cv}{foot}", "=id_periode", f(8, color="808080"), None, RIGHT)

    last = start_row + 4 * 8
    put(ws, f"B{last}", "Potong mengikuti garis kotak. Kartu kosong berarti nomor "
        "tersebut belum ada datanya.", f(9, italic=True, color="595959"), align=LEFT)
    ws.merge_cells(f"B{last}:H{last}")

    lock_sheet(ws)
    page_print(ws, area=f"B{start_row}:H{last - 1}", landscape=False)
    return ws


# ============================================================= 10. ABSENSI


def build_absensi(wb):
    ws = wb.create_sheet("ABSENSI")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [6, 30, 22, 15, 22, 3, 18, 16, 3, 3]):
        ws.column_dimensions[col].width = w

    gambar(ws, "logo-emblem.png", "A2", 80)
    put(ws, "A2", "DAFTAR HADIR KEGIATAN", f(16, True), align=CENTER)
    ws.merge_cells("A2:E2")
    put(ws, "A3", '=id_team&"  —  "&id_calon', f(12, True), align=CENTER)
    ws.merge_cells("A3:E3")
    put(ws, "A4", '="DESA "&id_desa', f(10), align=CENTER)
    ws.merge_cells("A4:E4")
    for col in "ABCDE":
        ws[f"{col}5"].border = Border(bottom=MED)

    put(ws, "G6", "PANEL — TIDAK DICETAK", f(9, True, WHITE), BLUE, CENTER)
    ws.merge_cells("G6:H6")
    fields = [("G7", "Nama kegiatan", "H7", None),
              ("G8", "Tanggal", "H8", None),
              ("G9", "Tempat", "H9", None),
              ("G10", "Pilih Kampung", "H10", "=daftar_kampung"),
              ("G11", "Pilih RT", "H11", "=daftar_rt"),
              ("G12", "Pilih TPS", "H12", "=daftar_tps")]
    for lab_ref, label, in_ref, formula in fields:
        put(ws, lab_ref, label, f(10, True), GREY_BG, LEFT, BOX)
        put(ws, in_ref, None, f(10, True, NAVY), INPUT, LEFT, BOX, "@", unlock=True)
        if formula:
            d = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(d)
            d.add(in_ref)
    put(ws, "G13", "Halaman ke-", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "H13", 1, f(11, True, NAVY), INPUT, CENTER, BOX, unlock=True)
    dvh = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1",
                         showErrorMessage=True, errorStyle="warning",
                         errorTitle="Halaman", error="Isi angka 1 atau lebih.")
    ws.add_data_validation(dvh)
    dvh.add("H13")
    put(ws, "G14", "Jumlah anggota", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "H14", f'=COUNT({DB}!$AD${FIRST}:$AD${LAST})', f(11, True, RED),
        None, CENTER, BOX, "#,##0")
    put(ws, "G15", "Total halaman", f(10, True), GREY_BG, LEFT, BOX)
    put(ws, "H15", f'=MAX(1,ROUNDUP($H$14/{N_ABSEN},0))', f(11, True, NAVY),
        None, CENTER, BOX, "#,##0")
    link(ws, "G17", "◀  MENU", "MENU!A1")
    for i, teks in enumerate([
            "Kosongkan filter Kadus/RT/TPS = daftar hadir SELURUH anggota.",
            "Isi satu filter untuk per Kadus, per RT, atau per TPS;",
            "isi Kadus + RT sekaligus untuk per korwil.",
            f"Satu halaman memuat {N_ABSEN} nama — ganti Halaman ke- lalu Ctrl+P."]):
        put(ws, f"G{18 + i}", teks, f(9, italic=True, color="595959"), align=LEFT)

    kelompok = ('=IF(COUNTA($H$10:$H$12)=0,"SELURUH ANGGOTA",'
                'MID(IF($H$10="",""," • "&$H$10)'
                '&IF($H$11="",""," • RT "&$H$11)'
                '&IF($H$12="",""," • TPS "&$H$12),4,200))')
    info = [("A7", "KEGIATAN", '=IF($H$7="","......................................",$H$7)'),
            ("A8", "HARI / TANGGAL", '=IF($H$8="","......................................",$H$8)'),
            ("A9", "TEMPAT", '=IF($H$9="","......................................",$H$9)'),
            ("A10", "KELOMPOK", kelompok),
            ("A11", "JUMLAH", '=$H$14&" orang"&IF($H$15>1,"     (halaman "&$H$13'
                              '&" dari "&$H$15&")","")')]
    for ref, label, formula in info:
        row = ref[1:]
        put(ws, ref, label, f(10, True), align=LEFT)
        ws.merge_cells(f"A{row}:B{row}")
        put(ws, f"C{row}", ":", f(10, True), align=CENTER)
        put(ws, f"D{row}", formula, f(10, True, NAVY), align=LEFT)
        ws.merge_cells(f"D{row}:E{row}")

    ws.column_dimensions["J"].hidden = True     # penunjuk baris DATABASE
    hd = 13
    header_row(ws, hd, [("NO", None), ("NAMA LENGKAP", None), ("JABATAN", None),
                        ("NO. HP", None), ("TANDA TANGAN", None)])
    src = {"B": "B", "C": "N", "D": "O"}
    for i in range(N_ABSEN):
        r = hd + 1 + i
        urut = f'(($H$13-1)*{N_ABSEN}+{i + 1})'
        put(ws, f"J{r}", f'=IFERROR(MATCH({urut},{DB}!$AD${FIRST}:$AD${LAST},0),"")',
            f(9))
        put(ws, f"A{r}", f'=IF($J{r}="","",{urut})', f(10), None, CENTER, BOX)
        for col, dbcol in src.items():
            put(ws, f"{col}{r}",
                f'=IF($J{r}="","",INDEX({DB}!${dbcol}${FIRST}:${dbcol}${LAST},$J{r})&"")',
                f(10), None, LEFT if col in ("B", "C") else CENTER, BOX,
                "@" if col == "D" else None)
        put(ws, f"E{r}", f'=IF($J{r}="","",$A{r}&".")', f(9, color="808080"),
            None, LEFT, BOX)
        ws.row_dimensions[r].height = 22

    ttd = hd + N_ABSEN + 2
    put(ws, f"D{ttd}", f'=id_desa&", ......................... "&id_tahun', f(11), align=CENTER)
    put(ws, f"D{ttd + 1}", "=id_jabatan_ttd", f(11), align=CENTER)
    put(ws, f"D{ttd + 5}", "=id_ketua", f(11, True), align=CENTER)
    for r in (ttd, ttd + 1, ttd + 5):
        ws.merge_cells(f"D{r}:E{r}")
    for col in "DE":
        ws[f"{col}{ttd + 5}"].border = Border(top=Side(style="thin", color="000000"))

    lock_sheet(ws)
    page_print(ws, area=f"A1:E{ttd + 6}", landscape=False, fit_height=1)
    return ws


# ============================================================== 11. PROFIL


def build_profil(wb):
    ws = wb.create_sheet("PROFIL")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [2, 12, 12, 12, 12, 12, 4, 34, 3]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "PROFIL & MEDIA KAMPANYE", width="H")
    put(ws, "C2", "Poster dan emblem resmi — siap dicetak atau dibagikan.",
        f(9, italic=True, color="595959"), align=LEFT)

    section(ws, 4, "POSTER RESMI", width="F")
    gambar(ws, "poster.jpg", "B5", 560)
    for r in range(5, 48):
        ws.row_dimensions[r].height = 15

    # emblem 200 px ≈ 10 baris (tinggi baris 15 pt = 20 px), keterangan
    # ditaruh tepat di bawahnya
    put(ws, "H4", "▌  EMBLEM RESMI", f(11, True, WHITE), NAVY, LEFT)
    gambar(ws, "logo-emblem.png", "H5", 200)
    put(ws, "H16", "Emblem utama — dipakai pada dasbor MENU, kop formulir "
        "CETAK dan ABSENSI, serta kartu anggota.",
        f(9, color="595959"), align=WRAP)
    ws.merge_cells("H16:H18")

    gambar(ws, "logo-emblem-alt.png", "H20", 200)
    put(ws, "H31", "Emblem alternatif tanpa pita nama — untuk spanduk, kaus, "
        "atau media yang butuh versi lebih ringkas.",
        f(9, color="595959"), align=WRAP)
    ws.merge_cells("H31:H33")

    put(ws, "H36", "Mengganti gambar: klik kanan pada gambar > Change Picture "
        "(Ganti Gambar). Ukuran dan posisinya tetap.",
        f(9, italic=True, color="595959"), align=WRAP)
    ws.merge_cells("H36:H38")

    put(ws, "B49", "Catatan ejaan: poster menuliskan JONI FAHAMSYAH, sedangkan "
        "pita pada emblem tertulis JONI FAHAMASYAH. Samakan lebih dulu bila "
        "keduanya dipakai berdampingan.", f(9, italic=True, color=RED), align=WRAP)
    ws.merge_cells("B49:F51")

    lock_sheet(ws)
    page_print(ws, area="A1:F48", landscape=False)
    return ws


# ============================================================ 12. PETUNJUK

PETUNJUK_ISI = [
    ("H", "A.  ALUR KERJA SINGKAT"),
    ("T", "1.  Buka sheet PENGATURAN. Sesuaikan nama Kadus, daftar RT, TPS, jabatan,"),
    ("T", "     angka TARGET tiap Kadus, serta identitas (nama calon, desa, ketua team)."),
    ("T", "2.  Buka sheet DATABASE. Ketik data anggota mulai baris kosong pertama."),
    ("T", "3.  Kolom NO, KORWIL, USIA, dan KELOMPOK USIA terisi sendiri (berlatar abu-abu)."),
    ("T", "4.  Buka sheet VALIDASI, perbaiki semua baris yang muncul di sana."),
    ("T", "5.  Pantau perkembangan di sheet REKAP dan TARGET."),
    ("T", "6.  Cetak lewat sheet CETAK (formulir), KARTU (kartu anggota), atau ABSENSI."),
    ("T", "7.  Sheet PROFIL memuat poster dan emblem resmi yang siap dicetak."),
    ("", ""),
    ("H", "B.  LOGO & KOP SURAT"),
    ("T", "Emblem tampil di dasbor MENU, kop sheet CETAK dan ABSENSI, serta di setiap"),
    ("T", "kartu pada sheet KARTU. Semuanya memakai berkas gambar yang sama."),
    ("T", "Mengganti logo: klik kanan pada gambar > Change Picture (Ganti Gambar);"),
    ("T", "ukuran dan posisinya tetap, jadi tata letak tidak perlu diatur ulang."),
    ("T", "Nama desa dan nama calon pada kop diambil dari sheet PENGATURAN, bukan"),
    ("T", "dari gambar — ubah di sana bila berganti wilayah atau periode."),
    ("", ""),
    ("H", "C.  MENCETAK — SEMUA, PER KADUS, PER RT, PER KORWIL, PER TPS"),
    ("T", "Sheet CETAK punya lima filter di panel kanan: Kadus, RT, TPS, jabatan, status."),
    ("T", "Filter yang dikosongkan berarti SEMUA, jadi satu sheet melayani semua kebutuhan:"),
    ("T", "   •  semua filter kosong          →  seluruh anggota"),
    ("T", "   •  isi Kampung saja             →  per kampung"),
    ("T", "   •  isi RT saja                  →  per RT"),
    ("T", "   •  isi Kampung + RT             →  per korwil"),
    ("T", "   •  isi TPS saja                 →  per TPS"),
    ("T", "Judul dan jumlah pada kop ikut berubah sendiri. Bila Total halaman lebih dari 1,"),
    ("T", "ganti angka Halaman ke- lalu Ctrl+P lagi. Sheet ABSENSI memakai cara yang sama,"),
    ("T", "sheet KARTU bisa disaring per Kadus/RT/TPS."),
    ("T", "Daftar sangat panjang dalam sekali cetak: buka sheet DATABASE, klik panah pada"),
    ("T", "judul kolom (TPS, RT, KADUS, KORWIL) untuk menyaring, lalu Ctrl+P. Baris yang"),
    ("T", "disembunyikan filter tidak ikut tercetak, jadi tidak ada halaman kosong."),
    ("", ""),
    ("H", "D.  ARTI WARNA"),
    ("T", "Kuning   : sel isian — boleh diketik/dipilih."),
    ("T", "Abu-abu  : hasil rumus otomatis — jangan diketik manual."),
    ("T", "Merah muda pada DATABASE : NIK ganda (dobel) — periksa dan perbaiki."),
    ("T", "Kuning muda pada DATABASE: NIK bukan 16 digit."),
    ("T", "Abu-abu miring pada DATABASE: anggota berstatus Nonaktif."),
    ("", ""),
    ("H", "E.  KOLOM PADA SHEET DATABASE"),
    ("T", "Identitas : NAMA LENGKAP, NIK KTP (16 digit), NO. KK (16 digit), L/P."),
    ("T", "Alamat    : KAMPUNG, RT, RW, ALAMAT (keterangan tambahan)."),
    ("T", "Pengurus  : NAMA RT, NAMA RK, NAMA KADUS — lihat daftar di sheet PENGATURAN."),
    ("T", "Lainnya   : TPS, JABATAN, NO. HP, TGL LAHIR, STATUS, TGL GABUNG, PEREKRUT."),
    ("T", "KORWIL, USIA, dan KELOMPOK USIA terisi otomatis — jangan diketik."),
    ("T", "NIK KTP milik perorangan dan tidak boleh kembar; NO. KK milik satu keluarga"),
    ("T", "sehingga beberapa anggota boleh punya nomor yang sama."),
    ("", ""),
    ("H", "F.  STRUKTUR PENGURUS — KENAPA MEMAKAI NAMA, BUKAN NOMOR"),
    ("T", "Nomor RT berulang di tiap kadus: RT 001 ada di Kadus 1, 2, maupun 3."),
    ("T", "Kampung Gamprit pun dipakai Kadus 2 dan Kadus 3 sekaligus."),
    ("T", "Karena itu nama pengurusnyalah yang membedakan satu RT dari RT lain."),
    ("T", "Daftar lengkapnya ada di sheet PENGATURAN, di bawah judul STRUKTUR PENGURUS,"),
    ("T", "berisi kadus, ketua RK, ketua RT, dan kampung masing-masing."),
    ("", ""),
    ("H", "G.  MENYIMPAN & CADANGAN"),
    ("T", "Tekan Ctrl+S setiap selesai menginput — seluruh sheet tersimpan sekaligus."),
    ("T", "Buat salinan cadangan tiap minggu: File > Save As, beri nama + tanggal."),
    ("T", "Simpan juga satu salinan di HP atau flashdisk sebagai cadangan kedua."),
    ("", ""),
    ("H", "H.  SHEET TERKUNCI"),
    ("T", "Sheet REKAP, TARGET, VALIDASI, CARI, CETAK, KARTU, dan ABSENSI dikunci agar"),
    ("T", "rumus tidak terhapus. Sel kuning tetap bisa diisi. Bila perlu mengubah isinya:"),
    ("T", "tab Review (Tinjau) > Unprotect Sheet — tanpa kata sandi."),
    ("T", "Sheet DATABASE dan PENGATURAN sengaja tidak dikunci agar bebas disunting."),
    ("", ""),
    ("H", "I.  KAPASITAS"),
    ("T", "Kapasitas 2.000 baris anggota, 40 pilihan Kadus/RT/jabatan/TPS,"),
    ("T", "20 Kadus dan 20 RT pada matriks REKAP, 150 baris hasil pada CARI dan VALIDASI,"),
    ("T", "25 nama per halaman pada CETAK dan ABSENSI, 8 kartu per halaman pada KARTU."),
    ("T", "Sisa kapasitas terlihat di sheet MENU dan pojok kanan atas sheet DATABASE."),
    ("", ""),
    ("H", "J.  MENAMBAH TOMBOL MAKRO — OPSIONAL"),
    ("T", "Tombol pada aplikasi ini memakai hyperlink sehingga aman dibuka di HP maupun"),
    ("T", "komputer tanpa peringatan keamanan. Bila ingin tombol SIMPAN dan FILTER otomatis:"),
    ("T", "1.  File > Save As > pilih tipe Excel Macro-Enabled Workbook (*.xlsm)."),
    ("T", "2.  Tekan Alt+F11 > menu Insert > Module > tempel kode di bawah ini."),
    ("T", "3.  Kembali ke Excel > tab Developer > Insert > Button (Form Control),"),
    ("T", "     gambar tombol pada sheet MENU, lalu pilih nama makronya."),
    ("", ""),
    ("C", "Sub SimpanData()"),
    ("C", "    ThisWorkbook.Save"),
    ("C", '    MsgBox "Data berhasil disimpan.", vbInformation, "Team Pemenangan"'),
    ("C", "End Sub"),
    ("", ""),
    ("C", "Sub TampilkanPerKorwil()"),
    ("C", "    Dim k As String"),
    ("C", '    k = InputBox("Ketik KORWIL, contoh: Kadus I - RT 001")'),
    ("C", '    If k = "" Then Exit Sub'),
    ("C", '    Sheets("DATABASE").Activate'),
    ("C", '    Sheets("DATABASE").ListObjects("tblAnggota").Range.AutoFilter _'),
    ("C", "        Field:=7, Criteria1:=k"),
    ("C", "End Sub"),
    ("", ""),
    ("C", "Sub TampilkanSemua()"),
    ("C", '    Sheets("DATABASE").ListObjects("tblAnggota").Range.AutoFilter Field:=7'),
    ("C", "End Sub"),
]

CONTOH_ISIAN = [
    ("NAMA LENGKAP", "BUDI SANTOSO", "huruf kapital, sesuai KTP"),
    ("NIK (16 digit)", "3201011203800001", "16 angka, tanpa spasi"),
    ("L/P", "L", "pilih L atau P dari dropdown"),
    ("KAMPUNG", "Kadus I", "pilih dari dropdown"),
    ("RT", "001", "pilih dari dropdown, 3 angka"),
    ("KORWIL", "Kadus I - RT 001", "terisi otomatis, jangan diketik"),
    ("JABATAN", "Koordinator Wilayah", "pilih dari dropdown"),
    ("NO. HP", "081234567890", "boleh diawali angka 0"),
    ("ALAMAT", "dekat masjid", "keterangan tambahan, boleh kosong"),
    ("TPS", "01", "pilih dari dropdown, 01 sampai 20"),
    ("TGL LAHIR", "12/03/1980", "format hh/bb/tttt"),
    ("USIA", "45", "terisi otomatis dari tanggal lahir"),
    ("KELOMPOK USIA", "Dewasa 36-50 th", "terisi otomatis"),
    ("STATUS", "Aktif", "Aktif / Calon / Nonaktif"),
    ("TGL GABUNG", "01/02/2026", "tanggal mulai bergabung"),
    ("PEREKRUT", "BUDI SANTOSO", "nama anggota yang merekrut"),
    ("CATATAN", "koordinator TPS 01", "catatan bebas"),
]


def build_petunjuk(wb):
    ws = wb.create_sheet("PETUNJUK")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [3, 4, 24, 26, 40, 3]):
        ws.column_dimensions[col].width = w

    title_block(ws, 1, "PETUNJUK PENGGUNAAN APLIKASI", width="E")

    row = 4
    for kind, text in PETUNJUK_ISI:
        if kind == "H":
            put(ws, f"B{row}", text, f(11, True, WHITE), NAVY, LEFT)
            ws.merge_cells(f"B{row}:E{row}")
            ws.row_dimensions[row].height = 19
        elif kind == "T":
            put(ws, f"C{row}", text, f(10), align=LEFT)
            ws.merge_cells(f"C{row}:E{row}")
        elif kind == "C":
            put(ws, f"B{row}", text, Font(name="Consolas", size=9, color="1F4E79"),
                GREY_BG, LEFT)
            ws.merge_cells(f"B{row}:E{row}")
        row += 1

    row += 1
    put(ws, f"B{row}", "K.  CONTOH FORMAT PENGISIAN SHEET DATABASE",
        f(11, True, WHITE), NAVY, LEFT)
    ws.merge_cells(f"B{row}:E{row}")
    row += 1
    for i, (a, b, c) in enumerate([("KOLOM", "CONTOH ISI", "KETERANGAN")] + CONTOH_ISIAN):
        r = row + i
        head = i == 0
        bg = BLUE if head else (BAND if i % 2 == 0 else WHITE)
        font = f(10, True, WHITE) if head else f(10)
        put(ws, f"C{r}", a, font, bg, LEFT, BOX)
        put(ws, f"D{r}", b, font, bg, LEFT, BOX, "@")
        put(ws, f"E{r}", c, font, bg, LEFT, BOX)
    end = row + len(CONTOH_ISIAN) + 1
    put(ws, f"C{end}", "Baris di atas hanya contoh format — bukan data anggota. "
        "Jangan disalin ke sheet DATABASE.", f(9, italic=True, color=RED), align=LEFT)
    ws.merge_cells(f"C{end}:E{end}")

    page_print(ws, area=f"A1:E{end}", landscape=False)
    return ws


# ======================================================== nama & penyusunan


def add_names(wb):
    names = {
        "daftar_kampung": f"OFFSET(REFERENSI!$B${REF_FIRST},0,0,"
                          f"MAX(1,COUNTA(REFERENSI!$B${REF_FIRST}:$B${REF_LAST})),1)",
        "daftar_rw": f"OFFSET(REFERENSI!$E${REF_FIRST},0,0,"
                     f"MAX(1,COUNTA(REFERENSI!$E${REF_FIRST}:$E${REF_LAST})),1)",
        "daftar_nama_rt": f"OFFSET(REFERENSI!$O${REF_FIRST},0,0,"
                          f"MAX(1,COUNTA(REFERENSI!$O${REF_FIRST}:$O${REF_LAST})),1)",
        "daftar_nama_rk": f"OFFSET(REFERENSI!$P${REF_FIRST},0,0,"
                          f"MAX(1,COUNTA(REFERENSI!$P${REF_FIRST}:$P${REF_LAST})),1)",
        "daftar_kadus": f"OFFSET(REFERENSI!$Q${REF_FIRST},0,0,"
                        f"MAX(1,COUNTA(REFERENSI!$Q${REF_FIRST}:$Q${REF_LAST})),1)",
        "daftar_rt": f"OFFSET(REFERENSI!$D${REF_FIRST},0,0,"
                     f"MAX(1,COUNTA(REFERENSI!$D${REF_FIRST}:$D${REF_LAST})),1)",
        "daftar_jabatan": f"OFFSET(REFERENSI!$F${REF_FIRST},0,0,"
                          f"MAX(1,COUNTA(REFERENSI!$F${REF_FIRST}:$F${REF_LAST})),1)",
        "daftar_tps": f"OFFSET(REFERENSI!$H${REF_FIRST},0,0,"
                      f"MAX(1,COUNTA(REFERENSI!$H${REF_FIRST}:$H${REF_LAST})),1)",
        "daftar_status": f"OFFSET(REFERENSI!$J${REF_FIRST},0,0,"
                         f"MAX(1,COUNTA(REFERENSI!$J${REF_FIRST}:$J${REF_FIRST + 4})),1)",
        "usia_label": f"REFERENSI!$L${REF_FIRST}:$L${REF_FIRST + len(USIA_DEFAULT) - 1}",
        "usia_min": f"REFERENSI!$M${REF_FIRST}:$M${REF_FIRST + len(USIA_DEFAULT) - 1}",
        "target_kadus": f"REFERENSI!$C${REF_FIRST}:$C${REF_LAST}",
        "nama_anggota": f"OFFSET({DB}!$B${FIRST},0,0,"
                        f"MAX(1,COUNTA({DB}!$B${FIRST}:$B${LAST})),1)",
        "id_calon": "REFERENSI!$C$49",
        "id_team": "REFERENSI!$C$50",
        "id_periode": "REFERENSI!$C$51",
        "id_tahun": "REFERENSI!$C$52",
        "id_desa": "REFERENSI!$C$53",
        "id_kecamatan": "REFERENSI!$C$54",
        "id_kabupaten": "REFERENSI!$C$55",
        "id_ketua": "REFERENSI!$C$56",
        "id_jabatan_ttd": "REFERENSI!$C$57",
    }
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


SAMPLE = [
    ("BUDI SANTOSO", "3201011203800001", "L", 0, 0, 0, "081234567801", "Dusun I RT 001"),
    ("SITI AMINAH", "3201014507850002", "P", 0, 0, 4, "081234567802", "Dusun I RT 001"),
    ("AGUS PRAYITNO", "3201010101900003", "L", 0, 1, 5, "081234567803", "Dusun I RT 002"),
    ("DEWI LESTARI", "3201015512920004", "P", 1, 2, 6, "081234567804", "Dusun II RT 003"),
    ("RUDI HARTONO", "3201012208880005", "L", 1, 2, 6, "081234567805", "Dusun II RT 003"),
    ("YULI ASTUTI", "3201016003950006", "P", 1, 3, 7, "081234567806", "Dusun II RT 004"),
    ("HENDRA GUNAWAN", "3201011711750007", "L", 2, 4, 4, "081234567807", "Dusun III RT 005"),
    ("NUR HALIMAH", "3201014201990008", "P", 2, 4, 6, "081234567808", "Dusun III RT 005"),
    ("JOKO SUSILO", "3201010506700009", "L", 3, 5, 1, "081234567809", "Dusun IV RT 006"),
    ("MAYA SARI", "3201015209010010", "P", 4, 6, 7, "081234567810", "Dusun V RT 007"),
]


def fill_sample(wb):
    """Isi data contoh untuk pengujian rumus (tidak dipakai pada berkas rilis)."""
    import datetime as dt
    ws = wb[DB]
    for i, (nama, nik, jk, ik, irt, ijab, hp, alamat) in enumerate(SAMPLE):
        r = FIRST + i
        ws[f"B{r}"] = nama
        ws[f"C{r}"] = nik
        ws[f"D{r}"] = jk
        ws[f"E{r}"] = KADUS_DEFAULT[ik]
        ws[f"F{r}"] = RT_DEFAULT[irt]
        ws[f"H{r}"] = JABATAN_DEFAULT[ijab]
        ws[f"I{r}"] = hp
        ws[f"J{r}"] = alamat
        ws[f"K{r}"] = TPS_DEFAULT[irt % len(TPS_DEFAULT)]
        ws[f"L{r}"] = dt.date(1970 + i * 3, 1 + i % 12, 1 + i % 27)
        ws[f"O{r}"] = "Aktif" if i % 4 else "Calon"
        ws[f"P{r}"] = dt.date(2026, 1, 1 + i)
    # satu NIK sengaja digandakan + satu NIK pendek untuk menguji VALIDASI
    ws[f"B{FIRST + 10}"] = "TEST NIK GANDA"
    ws[f"C{FIRST + 10}"] = SAMPLE[0][1]
    ws[f"E{FIRST + 10}"] = KADUS_DEFAULT[0]
    ws[f"F{FIRST + 10}"] = RT_DEFAULT[0]
    ws[f"H{FIRST + 10}"] = JABATAN_DEFAULT[6]
    ws[f"I{FIRST + 10}"] = "08999"
    ws[f"J{FIRST + 10}"] = "Dusun I RT 001"
    ws[f"B{FIRST + 11}"] = "TEST NIK PENDEK"
    ws[f"C{FIRST + 11}"] = "32010112038"
    ws[f"E{FIRST + 11}"] = KADUS_DEFAULT[1]
    ws[f"F{FIRST + 11}"] = RT_DEFAULT[2]
    ws[f"J{FIRST + 11}"] = "Dusun II RT 003"


def build(path, sample=False):
    wb = Workbook()
    wb.remove(wb.active)

    build_referensi(wb)
    build_database(wb)
    build_menu(wb)
    build_cari(wb)
    build_rekap(wb)
    build_target(wb)
    build_validasi(wb)
    build_cetak(wb)
    build_kartu(wb)
    build_absensi(wb)
    build_profil(wb)
    build_petunjuk(wb)
    add_names(wb)

    order = ["MENU", "DATABASE", "CARI", "REKAP", "TARGET", "VALIDASI",
             "CETAK", "KARTU", "ABSENSI", "PROFIL", "REFERENSI", "PETUNJUK"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0

    wb.properties.title = "Aplikasi Team Pemenangan"
    wb.properties.subject = "Database anggota team pemenangan tingkat desa"
    wb.properties.creator = "Aplikasi Team Pemenangan"
    wb.calculation.fullCalcOnLoad = True

    if sample:
        fill_sample(wb)

    wb.save(path)
    return path


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    out = args[0] if args else "Aplikasi_Team_Pemenangan.xlsx"
    build(out, sample="--sample" in sys.argv)
    print(f"tersimpan: {out}")
