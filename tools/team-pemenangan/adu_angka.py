#!/usr/bin/env python3
"""Adu setiap angka hasil hitungan Excel dengan hitungan aplikasi HTML.

Keduanya menghitung hal yang sama dari data yang sama, jadi kalau berbeda salah
satunya salah. Yang penting: perbedaan seperti itu TIDAK muncul sebagai galat
rumus — berkasnya tetap dilaporkan "nol galat".

Alat ini lahir setelah dua kali kejadian. Pertama, mengubah arti kolom indeks
membuat seluruh jumlah hasil saringan terbaca 10.000; hitung-ulang tetap
melaporkan nol galat. Kedua, blok gender dan status pada REKAP ternyata sudah
lama menghitung kolom yang salah — NO. KK dan NO. HP, bukan L/P dan STATUS —
dan tidak pernah terlihat karena kedua kolom itu memang belum terisi pada data
awal, sehingga hasilnya nol dan nol pun tampak wajar.

Karena itu satu hal yang harus diingat saat memakainya: JANGAN hanya menguji
dengan data awal. Kolom yang kosong menyembunyikan kesalahan. Berikan data yang
kolom-kolomnya terisi bervariasi.

Jalankan:
    python3 adu_angka.py <folder-kerja> [database.json]

Keluar dengan status 1 bila ada angka yang berbeda, jadi bisa dipakai sebagai
penjaga sebelum berkas dibagikan.
"""
import asyncio, json, pathlib, subprocess, sys, warnings
import openpyxl
from playwright.async_api import async_playwright
warnings.filterwarnings('ignore')

BASE = pathlib.Path(__file__).resolve().parent
HTML = BASE / 'web' / 'Aplikasi_Team_Pemenangan.html'
S = pathlib.Path(sys.argv[1])
XLSX = S / 'adu.xlsx'
DATA = pathlib.Path(sys.argv[2]) if len(sys.argv) > 2 else BASE / 'web/seed/database.json'

subprocess.run([sys.executable, 'build_workbook.py', str(XLSX)], cwd=BASE, check=True, capture_output=True)
subprocess.run([sys.executable, 'isi_data.py', str(XLSX), str(DATA)], cwd=BASE, check=True, capture_output=True)
r = subprocess.run([sys.executable, str(pathlib.Path.home()/'.claude/skills/xlsx/scripts/recalc.py'),
                    str(XLSX), '480'], capture_output=True, text=True)
print('recalc:', json.loads(r.stdout)['status'], '|',
      json.loads(r.stdout).get('total_errors'), 'galat')

wb = openpyxl.load_workbook(XLSX, data_only=True)
def sel(sheet, ref): return wb[sheet][ref].value

def cari_label(sheet, kol_label, kol_nilai, teks, sampai=40):
    ws = wb[sheet]
    for r in range(1, sampai):
        v = ws[f'{kol_label}{r}'].value
        if v and teks.lower() in str(v).lower():
            return ws[f'{kol_nilai}{r}'].value
    return '(tidak ketemu)'

seed_ms = json.loads(DATA.read_text())['members']

async def main():
    seed = json.loads(DATA.read_text())
    async with async_playwright() as p:
        b = await p.chromium.launch(
            executable_path='/opt/pw-browsers/chromium-1194/chrome-linux/chrome',
            args=['--no-sandbox'])
        pg = await b.new_page()
        await pg.goto(HTML.as_uri()); await pg.wait_for_timeout(300)
        await pg.evaluate("""(d) => localStorage.setItem('tp_app_v1', JSON.stringify(
            {v:3, disimpan:d.disimpan, members:d.members, settings:d.settings}))""", seed)
        await pg.reload(); await pg.wait_for_timeout(900)
        js = await pg.evaluate("""() => {
            const r = ringkasan();
            const perKp = {}; perKadus().forEach(x => { perKp[x.kadus] = x.jumlah; });
            const perRt = {}; hitungPer('rt').forEach((v, k) => { perRt[k] = v; });
            const perTps = {}; hitungPer('tps').forEach((v, k) => { perTps[k] = v; });
            const perJab = {}; hitungPer('jabatan').forEach((v, k) => { perJab[k] = v; });
            const perKorwil = {}; hitungPer('korwil').forEach((v, k) => { perKorwil[k] = v; });
            const nikC = hitungNik();
            const mas = {};
            state.members.forEach(m => { const x = masalah(m, nikC); if (x) mas[x] = (mas[x]||0)+1; });
            const usiaG = {};
            state.settings.kelompokUsia.forEach(g => { usiaG[g.label] = 0; });
            state.members.forEach(m => { const g = kelompokUsia(usia(m)); if (g) usiaG[g]++; });
            return {total: r.total, aktif: r.aktif, L: r.L, P: r.P,
                    kampungAktif: r.kadusAktif, rtAktif: r.rtAktif, tpsAktif: r.tpsAktif,
                    bermasalah: r.bermasalah, perKp, perRt, perTps, perJab, perKorwil,
                    mas, usiaG};
        }""")
        await b.close()
    return js

js = asyncio.run(main())
beda = []
def adu(nama, xl, html):
    sama = (xl or 0) == (html or 0)
    if not sama: beda.append((nama, xl, html))
    print(f"  {'OK ' if sama else 'BEDA'}  {nama:38} Excel={xl!r:>12}  HTML={html!r:>12}")

print('\n--- MENU')
adu('Total anggota', cari_label('MENU','E','F','Total anggota'), js['total'])
adu('Berstatus Aktif', cari_label('MENU','E','F','Aktif'), js['aktif'])
adu('Kampung aktif', cari_label('MENU','E','F','kampung aktif'), js['kampungAktif'])
adu('RT tercakup', cari_label('MENU','E','F','RT tercakup'), js['rtAktif'])
adu('Perlu diperbaiki', cari_label('MENU','E','F','perlu diperbaiki'), js['bermasalah'])

print('\n--- REKAP per kampung (total)')
ws = wb['REKAP']
for r in range(6, 26):
    kp = ws[f'B{r}'].value
    if not kp or kp == 'TOTAL': continue
    adu(f'kampung {kp}', ws['W'+str(r)].value, js['perKp'].get(kp, 0))

print('\n--- REKAP matriks RT (kolom total per RT)')
hd = next(r for r in range(1, 40) if ws[f'B{r}'].value == 'KAMPUNG')
# baris TOTAL dicari lewat labelnya, bukan nomor tetap — tepat inilah yang
# sempat membuat pembanding ini salah membaca setelah satu baris disisipkan
baris_tot = next(r for r in range(hd, hd + 40) if ws[f'B{r}'].value == 'TOTAL')
baris_sisa = next(r for r in range(hd, baris_tot)
                  if str(ws[f'B{r}'].value or '').startswith('(kampung'))
print(f'  (baris judul {hd}, "(kampung belum diisi)" {baris_sisa}, TOTAL {baris_tot})')
for c in range(3, 23):
    rt = ws.cell(hd, c).value
    if not rt: continue
    tot = ws.cell(baris_tot, c).value
    if tot is None: continue
    adu(f'RT {rt}', tot, js['perRt'].get(str(rt), 0))
print('  --- baris "(kampung belum diisi)" harus menutup selisihnya')
adu('tanpa kampung (total)', ws.cell(baris_sisa, 3 + 20).value,
    sum(1 for m in seed_ms if not m['kampung'] and m['nama'].strip()))

print('\n--- REKAP gender & status per kampung')
hb = next(r for r in range(1, 80) if ws[f'B{r}'].value == 'KAMPUNG'
          and ws[f'C{r}'].value == 'LAKI-LAKI')
tot_b = next(r for r in range(hb, hb + 40) if ws[f'B{r}'].value == 'TOTAL')
sisa_b = next(r for r in range(hb, tot_b)
              if str(ws[f'B{r}'].value or '').startswith('(kampung'))
def hit(f):
    return sum(1 for m in seed_ms if m['nama'].strip() and f(m))
adu('L (total)', ws.cell(tot_b, 3).value, hit(lambda m: m['jk'] == 'L'))
adu('P (total)', ws.cell(tot_b, 4).value, hit(lambda m: m['jk'] == 'P'))
adu('Aktif (total)', ws.cell(tot_b, 5).value, hit(lambda m: m['status'] == 'Aktif'))
adu('Calon (total)', ws.cell(tot_b, 6).value, hit(lambda m: m['status'] == 'Calon'))
adu('Nonaktif (total)', ws.cell(tot_b, 7).value, hit(lambda m: m['status'] == 'Nonaktif'))
adu('Jumlah orang (total)', ws.cell(tot_b, 8).value, hit(lambda m: True))
adu('tanpa kampung (gender)', ws.cell(sisa_b, 8).value, hit(lambda m: not m['kampung']))

print('\n--- TARGET realisasi per kampung')
tg = wb['TARGET']
hdt = next(r for r in range(1, 30) if str(tg[f'B{r}'].value or '').upper() in ('KADUS', 'KAMPUNG'))
for r in range(hdt + 1, hdt + 1 + 24):
    kp = tg[f'B{r}'].value
    if not kp or kp == 'TOTAL': continue
    adu(f'realisasi {kp}', tg[f'D{r}'].value, js['perKp'].get(kp, 0))

print('\n--- VALIDASI ringkasan masalah')
v = wb['VALIDASI']
peta = {'NIK kosong':'NIK kosong', 'NIK bukan 16 digit':'NIK bukan 16 digit',
        'NIK ganda':'NIK ganda', 'Jabatan kosong':'Jabatan kosong',
        'No. HP kosong':'No. HP kosong', 'Alamat kosong':'Alamat kosong',
        'Usia di bawah 17':'Usia di bawah 17'}
for r in range(5, 22):
    lab = v[f'B{r}'].value
    if not lab: continue
    for kunci, hkey in peta.items():
        if str(lab).lower().startswith(kunci.lower()[:12]):
            adu(f'masalah: {kunci}', v[f'E{r}'].value, js['mas'].get(hkey, 0))
            break
adu('total bermasalah', cari_label('VALIDASI','B','E','Total baris bermasalah'), js['bermasalah'])

print(f"\n{'='*66}")
print('BEDA:', len(beda) if beda else 'tidak ada — seluruh angka cocok')
for n, a, b in beda: print(f'   {n}: Excel={a!r} HTML={b!r}')
sys.exit(1 if beda else 0)
